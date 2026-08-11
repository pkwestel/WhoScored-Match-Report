"""
Combined WhoScored + FotMob Match Report
==========================================

Merges the two independently-scraped reports (whoscored_report.py and
fotmob_report.py) for the SAME match into one workbook: every tab from
each source is kept exactly as that source's own build_workbook() already
produces it, EXCEPT each source's own "Notes" tab - those are dropped
entirely (along with this module's own former "Notes" tab), by request;
none of "WS - Notes" / "FM - Notes" / "Notes" appear in this workbook.
Totals is the only remaining name that collides between the sources, so
it's the only tab still prefixed ("WS - " / "FM - "); every other tab
keeps its own plain name - see build_combined_workbook() below.

The one genuinely new thing here is the "Shot Creating Actions" tab -
WhoScored's own shot-creating-actions table (Player, Distance, Body Part,
SCA1/SCA2) with FotMob's own Minute/Added Time plus its shot-quality fields
(xG, PSxG, Outcome, Situation) attached to each row - see
compute_combined_shots() for the matching logic and its limitations. Split
into two tables (one per team), each followed by a small "Top 3 Shots by
xG" mini table for that team - see build_combined_workbook() below.

This module has no scraping of its own - combined_streamlit_app.py drives
whoscored_report.py and fotmob_report.py's own scrape/compute/build_workbook
functions directly, then hands the two resulting Workbook objects (plus the
combined shots table) to build_combined_workbook() here.
"""

import re
import unicodedata
from copy import copy

import pandas as pd
from openpyxl import Workbook

import whoscored_report as wr

# FotMob's "Hit Post" outcome is relabeled "Woodwork" in the combined Shot
# Creating Actions tab specifically (per request) - this does NOT change
# fotmob_report.py's own Shots/Shot Breakdown/Totals tabs, which keep
# FotMob's original "Hit Post" wording untouched.
OUTCOME_RELABEL = {'Hit Post': 'Woodwork'}

COMBINED_SHOTS_COLUMNS = [
    'Minute', 'Added Time', 'Player', 'Team', 'xG', 'PSxG', 'Outcome', 'Distance (yd)',
    'Body Part', 'Situation', 'SCA1_Player', 'SCA1_Action', 'SCA2_Player', 'SCA2_Action',
]

# WhoScored and FotMob don't always use the same name for the same club
# (e.g. one may say "Man Utd", the other "Manchester United"; "Spurs" vs
# "Tottenham Hotspur") - plain accent/case normalization alone can't catch
# that, so this maps known alternate/short forms to one canonical name.
# Every value in this dict is a list of alternates for that canonical name -
# add more entries here if a new club turns up with a mismatch between the
# two sites; it's a plain dict, no other code needs to change.
TEAM_NAME_ALIASES = {
    'Manchester United': ['Man Utd', 'Man United', 'Man Utd.', 'Man U'],
    'Manchester City': ['Man City', 'Man City.'],
    'Tottenham Hotspur': ['Spurs', 'Tottenham'],
    'Newcastle United': ['Newcastle'],
    'Wolverhampton Wanderers': ['Wolves', 'Wolverhampton'],
    'Brighton & Hove Albion': ['Brighton', 'Brighton and Hove Albion', 'Brighton Hove Albion'],
    'West Ham United': ['West Ham'],
    'Nottingham Forest': ["Nott'm Forest", 'Notts Forest', 'Forest'],
    'Leicester City': ['Leicester'],
    'AFC Bournemouth': ['Bournemouth'],
    'Sheffield United': ['Sheffield Utd', 'Sheff Utd', 'Sheff United'],
    'West Bromwich Albion': ['West Brom', 'WBA'],
    'Leeds United': ['Leeds'],
    'Cardiff City': ['Cardiff'],
    'Norwich City': ['Norwich'],
    'Stoke City': ['Stoke'],
    'Swansea City': ['Swansea'],
    'Huddersfield Town': ['Huddersfield'],
    'Queens Park Rangers': ['QPR'],
    'Blackburn Rovers': ['Blackburn'],
    'Preston North End': ['Preston'],
    'Ipswich Town': ['Ipswich'],
    'Crystal Palace': ['Palace'],
    'Aston Villa': ['Villa'],
    'Southampton': ['Saints'],
    'Luton Town': ['Luton'],
}


def _normalize_name(name):
    """
    Strip accents/diacritics/punctuation and normalize case/whitespace -
    the first step of matching a name ACROSS WhoScored and FotMob (see
    canonical_team_name() below, which layers TEAM_NAME_ALIASES on top of this
    for the cases plain normalization can't fix, like "Spurs" vs
    "Tottenham Hotspur"). Never used for anything displayed.
    """
    if not isinstance(name, str):
        return ''
    nfkd = unicodedata.normalize('NFKD', name)
    ascii_name = ''.join(c for c in nfkd if not unicodedata.combining(c))
    ascii_name = re.sub(r'[^a-zA-Z0-9]+', ' ', ascii_name)
    return ' '.join(ascii_name.lower().split())


# Built once at import time: every canonical name AND every one of its
# aliases (all run through _normalize_name) point to the same normalized
# canonical name, so canonical_team_name() below is a single dict lookup.
_TEAM_CANONICAL_LOOKUP = {}
for _canonical, _aliases in TEAM_NAME_ALIASES.items():
    _canon_norm = _normalize_name(_canonical)
    _TEAM_CANONICAL_LOOKUP[_canon_norm] = _canon_norm
    for _alias in _aliases:
        _TEAM_CANONICAL_LOOKUP[_normalize_name(_alias)] = _canon_norm


def canonical_team_name(name):
    """
    Canonicalize a team name for cross-source matching in
    compute_combined_shots() - looks the normalized name up in
    TEAM_NAME_ALIASES; if it's not a known alias, falls back to the plain
    normalized name as-is (so two sources that already agree on a team's
    name still match normally, with no alias table entry needed).
    """
    norm = _normalize_name(name)
    return _TEAM_CANONICAL_LOOKUP.get(norm, norm)


def compute_combined_shots(sca_out, fm_shots_df):
    """
    Shot Creating Actions tab (combined report): one row per shot, using
    WhoScored's own shot list (sca_out, from whoscored_report.compute_sca -
    already excludes own goals) as the spine for Player/Team/Distance/Body
    Part/SCA1/SCA2, with FotMob's own Minute/Added Time and shot-quality
    fields (xG, PSxG/xGOT, Outcome, Situation) attached to each row.

    WhoScored and FotMob share no common shot ID, so shots are matched by
    TEAM, then paired up in strict CHRONOLOGICAL ORDER within that team -
    a team's 1st shot of the match on WhoScored is matched to that same
    team's 1st shot on FotMob, 2nd to 2nd, and so on (WhoScored's own
    'minute' and FotMob's Minute+Added Time are each used only to sort
    that source's own shots into order, not to match across sources
    directly - the two sites don't always log stoppage-time minutes
    identically, so matching on relative order within the team is more
    reliable than matching on the minute value itself). This assumes both
    sources recorded the same NUMBER of shots for that team - if one side
    has more (a scrape gap, most likely), the extra shot(s) at the end of
    that team's list are left with blank FotMob fields, since there's no
    way to tell which specific shot is the unmatched one from timing alone.

    Minute/Added Time in the output are FotMob's own fields (not
    WhoScored's minute) - once matched, FotMob's own stoppage-time
    bookkeeping is used for both columns.

    Outcome: FotMob's 'Hit Post' is relabeled 'Woodwork' in this column.
    """
    if sca_out is None or sca_out.empty:
        return pd.DataFrame(columns=COMBINED_SHOTS_COLUMNS)

    ws = sca_out.copy().reset_index(drop=True)
    ws['_norm_team'] = ws['team'].apply(canonical_team_name)
    ws['_orig_order'] = range(len(ws))

    ws_sorted = ws.sort_values(['_norm_team', 'minute']).copy()
    ws_sorted['_rank_in_team'] = ws_sorted.groupby('_norm_team').cumcount()

    fm_cols = ['_norm_team', '_rank_in_team', 'Minute', 'Added Time', 'xG', 'xGOT', 'Outcome', 'Situation']
    if fm_shots_df is not None and not fm_shots_df.empty:
        fm = fm_shots_df.copy()
        fm['_norm_team'] = fm['Team'].apply(canonical_team_name)
        fm['_effective_minute'] = fm['Minute'].fillna(0) + fm['Added Time'].fillna(0)
        fm = fm.sort_values(['_norm_team', '_effective_minute']).copy()
        fm['_rank_in_team'] = fm.groupby('_norm_team').cumcount()
        fm_for_merge = fm[fm_cols]
    else:
        fm_for_merge = pd.DataFrame(columns=fm_cols)

    merged = ws_sorted.merge(fm_for_merge, on=['_norm_team', '_rank_in_team'], how='left')
    merged = merged.sort_values('_orig_order').reset_index(drop=True)
    merged['Outcome'] = merged['Outcome'].apply(lambda o: OUTCOME_RELABEL.get(o, o))

    out = pd.DataFrame({
        'Minute': merged['Minute'],
        'Added Time': merged['Added Time'],
        'Player': merged['player'],
        'Team': merged['team'],
        'xG': merged['xG'],
        'PSxG': merged['xGOT'],
        'Outcome': merged['Outcome'],
        'Distance (yd)': merged['shot_distance_yd'],
        'Body Part': merged['bodyPart'],
        'Situation': merged['Situation'],
        'SCA1_Player': merged['sca1_player'],
        'SCA1_Action': merged['sca1_action'],
        'SCA2_Player': merged['sca2_player'],
        'SCA2_Action': merged['sca2_action'],
    })
    return out[COMBINED_SHOTS_COLUMNS]


def _copy_sheet(src_ws, dst_wb, new_title):
    """
    Copy a worksheet's values, styles, column widths, and freeze panes into
    a new sheet - possibly in a different Workbook. openpyxl's own
    Workbook.copy_worksheet() only works WITHIN the same workbook, so this
    fills that gap, letting build_combined_workbook() below reuse each
    source report's own already-built Workbook wholesale rather than
    re-implementing every tab's writing logic a second time.
    """
    dst_ws = dst_wb.create_sheet(new_title)
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
                new_cell.border = copy(cell.border)
                new_cell.number_format = cell.number_format
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes
    return dst_ws


def build_combined_workbook(wb_ws, wb_fm, combined_shots, home_name=None, away_name=None):
    """
    Assembles the combined workbook from two ALREADY-BUILT standalone
    workbooks (whoscored_report.build_workbook() / fotmob_report.
    build_workbook()) plus the new merged shots table. Every sheet from
    each source is copied over wholesale, EXCEPT: WhoScored's own
    "Shot Creating Actions" sheet and FotMob's own "Shots" sheet, which are
    dropped entirely since they're replaced by the single new merged
    "Shot Creating Actions" tab built here instead; and both sources' own
    "Notes" sheets, along with the combined report's own former "Notes"
    tab, which are dropped entirely and not rebuilt here - none of the
    three ("WS - Notes", "FM - Notes", "Notes") are included in this
    workbook, by request. Totals is the only remaining name that collides
    between the sources, so it's the only tab still prefixed "WS - " /
    "FM - "; every other tab keeps its own plain name.

    home_name/away_name (WhoScored's own team names, matching the Team
    values already on combined_shots - see compute_combined_shots) control
    which team's table is written first; if not supplied, teams are ordered
    alphabetically. The merged Shot Creating Actions tab is split into two
    tables, one per team, and each team's table is followed - two blank
    rows down - by a small "Top 3 Shots by xG" table (Minute, Player, xG)
    for that team, before moving on to the next team's section. Penalties
    (Situation == 'Penalty') are excluded from that mini table specifically -
    a penalty's xG is a fixed, known value rather than a reflection of shot
    quality, so it would otherwise crowd out more meaningful entries. The
    main per-team Shot Creating Actions table above it still includes
    penalties untouched.
    """
    wb = Workbook()
    wb.remove(wb.active)

    ws_prefixed_titles = {'Totals'}

    for title in wb_ws.sheetnames:
        if title == 'Notes':
            continue  # Notes tabs are dropped entirely from the combined report
        if title == 'Shot Creating Actions':
            # Replaced by the new merged tab, inserted right here so it
            # keeps the same relative position it held in the standalone
            # WhoScored workbook, rather than landing after every other
            # WS tab.
            ws = wb.create_sheet('Shot Creating Actions')
            teams_for_sca = ([t for t in [home_name, away_name] if t is not None]
                              or sorted(combined_shots['Team'].dropna().unique()))
            row = 1
            for t in teams_for_sca:
                t_shots = combined_shots[combined_shots['Team'] == t].drop(columns=['Team']).reset_index(drop=True)
                row = wr.write_table(ws, t_shots, start_row=row, title=f'{t} - Shot Creating Actions') + 3

                top3 = (t_shots[t_shots['Situation'] != 'Penalty'][['Minute', 'Player', 'xG']]
                        .dropna(subset=['xG'])
                        .sort_values('xG', ascending=False)
                        .head(3)
                        .reset_index(drop=True))
                row = wr.write_table(ws, top3, start_row=row, title=f'{t} - Top 3 Shots by xG') + 3
            wr.autosize(ws)
            ws.freeze_panes = 'A2'
            continue
        new_title = (f'WS - {title}' if title in ws_prefixed_titles else title)[:31]
        _copy_sheet(wb_ws[title], wb, new_title)

    for title in wb_fm.sheetnames:
        if title in ('Shots', 'Notes'):
            continue  # Shots is replaced by the new merged tab above; Notes tabs are dropped entirely
        new_title = (f'FM - {title}' if title in ws_prefixed_titles else title)[:31]
        _copy_sheet(wb_fm[title], wb, new_title)

    return wb
