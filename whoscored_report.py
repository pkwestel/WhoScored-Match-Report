"""
WhoScored match report generator.

Scrapes a WhoScored match-centre page and produces a styled Excel workbook
with tabs: Totals, Against, Touches, Passing, Shot Creating Actions,
Defensive Actions, Defensive Action Location, Passing Pairs, Shot Pairs, and
On-Off (plus a Notes tab documenting every definition/assumption used).

SETUP (one-time):
    Drop this file into the root of your cloned football-data-webscraping
    repo (the same folder that contains the `whoscored/` and `utils/`
    folders), with its virtual environment active and dependencies
    installed (`pip install -r requirements.txt`, plus `pip install openpyxl`).

USAGE:
    python whoscored_report.py "<whoscored match centre url>"

    Example:
    python whoscored_report.py "https://www.whoscored.com/matches/1903410/live/england-premier-league-2025-2026-manchester-united-crystal-palace"

    The output file (e.g. ManchesterUnited_vs_CrystalPalace.xlsx) is saved
    in the same folder you ran the script from.

NOTES / ASSUMPTIONS (also written into the workbook's "Notes" tab):
  - Pitch dimensions are assumed to be 105m x 68m (standard Opta convention)
    to convert the 0-100 normalized x/y coordinates into yards.
  - Progressive pass: a completed, OPEN PLAY pass that moves the ball at
    least 10 yards closer to the opponent's goal than the most-advanced
    point reached in that team's previous six completed passes (a rolling
    window that is NOT reset by turnovers), OR any completed, open play
    pass into the penalty area. Corners, free kicks, throw-ins, goal kicks,
    and keeper throws can never count as progressive (they still feed the
    rolling baseline). Passes starting in the defending 40% of the pitch
    are excluded, unless they end in the box.
  - Progressive carry: WhoScored/Opta data has no explicit "carry" event
    (unlike StatsBomb-sourced data), so carries are inferred using the same
    approach as well-known open-source WhoScored report notebooks: for each
    event, look ahead to the next one, skipping past TakeOns and Fouls
    (which don't represent a genuine loss of the ball), then treat the gap
    between the first event's end location and that next event's start
    location as a carry - provided both belong to the same team, the first
    event isn't a bare "BallTouch", the distance covered is between 3 and
    60 metres, and the elapsed time is between 1 and 10 seconds. A carry
    counts as progressive if it moves the ball at least 10 (pitch-length)
    units closer to the opponent's goal, or it enters the penalty area.
    Carries starting in a team's own third still count, but carries ENDING
    in their own third are excluded.
  - Carries into the final third / into the box: a carry that starts
    outside that zone and ends inside it (an entry, not merely occurring
    within it).
  - Shot-creating actions: the up-to-two offensive actions (passes,
    take-ons, defensive actions winning the ball, fouls drawn, or shots
    leading to a rebound shot) immediately before a shot, walking backward
    through the same team's play. Aerial duels and ground challenges never
    break this search (since they're recorded as separate rows for both
    sides and aren't genuine turnovers), but only a won *challenge* counts
    as a contributing action - a won aerial does not count, it's simply
    skipped over so the search can keep looking further back. The same
    applies to an opponent's Clearance, blocked pass, or keeper Save
    immediately before the shot - a defender's partial touch that doesn't
    fully clear the danger (e.g. heading a cross away for it to fall to
    another attacker) doesn't count as an SCA itself and doesn't break the
    search either, so the real originating pass still gets found and
    credited (this matters a lot for headers off crosses/corners).
    Rebounds are a special case: if the same team's own PREVIOUS shot is
    the action immediately before this one, that earlier shot IS the SCA
    (labeled just "Shot") and the search stops there - a rebound never
    gets a second SCA behind it. A Tackle or Interception winning the ball
    still counts as an SCA, but a BallRecovery (picking up a loose ball,
    not winning a contested one) does not - it's skipped over rather than
    counted, so the search keeps looking for the real contributing action.
  - Shot body part is read directly from WhoScored's own qualifiers on the
    shot event (Head / Left Foot / Right Foot / Other).
  - Pitch thirds: Own third = x < 33.3, Middle third = 33.3 <= x < 66.6,
    Final third = x >= 66.6 (x only, 0-100 normalized scale).
  - Attacking box = inside the opponent's 18-yard box (18 yards deep,
    44 yards wide, centered on goal).
  - Passing tab (per player, by team): Passes Completed/Attempted count all
    pass types; Passes Forward = attempts moving the ball's X coordinate
    forward regardless of success; Headed = pass attempts with a HeadPass
    qualifier; Crosses Attempted/Completed are open-play only (excludes
    corners and free kicks); Passes into Final 1/3 / into the Box are
    completed, open-play passes starting outside and ending inside that
    zone; Progressive Passes uses the same rolling-window definition
    described above; Shot Assists comes straight from WhoScored's own
    ShotAssist qualifier; SCA is the combined SCA1+SCA2 total from the Shot
    Creating Actions tab.
  - Totals tab (per team): rolls up the same numbers from the Touches,
    Passing, and Shot Creating Actions tabs - Shots is
    a count of shot events, Total Passes is the sum of each player's Passes
    Attempted, Total SCA is the sum of each player's SCA column.
  - 10+ Pass Sequences / Avg Passes per Sequence: possession sequences use
    a windowed "possession chain" algorithm (ported from the open-source
    notebooks behind insight90.streamlit.app) - possession only flips teams
    on a sustained run of opponent events within a sliding window, not on
    a single touch, and always restarts at goals/new periods. Tuned against
    insight90's published numbers for this match; close but not exact.
    "Passes" = all pass attempts per sequence (not just completed ones).
  - Field Tilt: each team's share of the two teams' combined final-third
    touches, as a percentage.
  - PPDA: opponent's passes / own defensive actions (Tackle/Interception/
    Foul/Challenge), both excluding each team's own deepest 30% of the
    pitch. Lower = more aggressive pressing. Tuned against a benchmark;
    close but not exact.
  - Tackles/Successful Tackles/Interceptions/Blocked Passes/Blocked Shots:
    exact counts from discrete WhoScored event types/qualifiers - verified
    to match a published benchmark exactly.
  - Defensive Action Height (m): MEAN x position of Tackle/Interception/
    Clearance/BallRecovery/Challenge events (goalkeeper excluded), converted
    to metres. Tuned against two real matches via diagnose_def_action_height.py;
    close but not exact (unlike the counts above).
  - Passes Received / Progressive Passes Received (Touches tab): receiver
    inferred as the next event's player, if that event is the same team.
  - Defensive Actions tab (per player, by team): Tackles/Interceptions are
    direct counts; Passes Blocked uses the BlockedPass event's own player;
    Shots Blocked uses the player of the 'Save' event immediately after a
    shot with a 'Blocked' qualifier (no direct player field exists for
    shot blocks) - verified to sum to the Totals tab's Blocked Shots.
  - Defensive Action Location tab (per player, by team): Tackles/
    Interceptions/Passes Blocked/Ball Recoveries, each split by pitch
    third (Own/Middle/Final) using the same third() rule as Touches.
  - Possession % (Totals tab): passes-based proxy - team's passes attempted
    / both teams' combined passes attempted. Close to a published benchmark.
  - Corners (Totals tab): count of Pass events with a 'CornerTaken'
    qualifier. Exact match against a published benchmark.
  - On-Off tab (per player, by team): Shots/Total Touches/thirds/Attacking
    Box/Carries into Final Third/Carries into Box/Passes into Final 1/3/
    Passes into the Box/Crosses Attempted/Crosses Completed (open play
    only), grouped as a block of all "For" (own team) columns followed by
    a block of all "Against" (opponent) columns, counted only during each
    player's own on-pitch window (kickoff or their SubstitutionOn time, to
    the final whistle or their own SubstitutionOff time). A red/second-
    yellow card isn't accounted for - a sent-off player's window still
    runs to the final whistle.

If WhoScored changes its page structure, the scrape step may need
adjusting - see `scrape_match()` below.
"""

import sys
import os
import re
import json
import ast
from collections import deque

import pandas as pd
import numpy as np
from bs4 import BeautifulSoup, NavigableString
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from utils.driver import get_driver

# ============================================================
# CONFIGURATION / ASSUMPTIONS
# ============================================================
PITCH_LEN_M = 105.0
PITCH_WID_M = 68.0
M_TO_YD = 1.09361

PROGRESSIVE_YARD_THRESHOLD = 10
PROGRESSIVE_ROLLING_WINDOW = 6
DEFENSIVE_START_CUTOFF = 40  # % of pitch (x < this is "defending 40%"), for passes

BOX_DEPTH_YD = 18
BOX_WIDTH_YD = 44

OWN_THIRD_MAX = 33.3
MIDDLE_THIRD_MAX = 66.6

OFFENSIVE_TYPES = {'Pass', 'TakeOn', 'Tackle', 'Interception',
                    'Clearance', 'Goal', 'SavedShot', 'MissedShots', 'ShotOnPost'}
# BallRecovery is deliberately excluded: picking up a loose ball isn't itself
# a shot-creating action, but it also shouldn't break the search - the walk-
# back just skips over it (same-team events not in OFFENSIVE_TYPES never
# break the search, only opponent-team events can) to find the real action
# that created the chance.
BOUNDARY_TYPES = {'Start', 'End', 'FormationSet', 'SubstitutionOn', 'SubstitutionOff'}
NON_BREAKING_DUEL_TYPES = {'Aerial', 'Challenge'}
COUNTABLE_DUEL_TYPES = {'Challenge'}  # Aerial never counts as an SCA, per user instruction
# Opponent-side events that are a loose-ball/deflection, not a clean turnover -
# a defender's partial clearance, a blocked shot/pass, or a keeper save can
# all put the ball right back to the shooting team, so the search should skip
# past them (not counting them as an SCA, but not breaking the chain either)
# to find the real originating pass. A genuine opponent Tackle/Interception/
# BallRecovery still breaks the chain - those represent an actual regain.
LOOSE_BALL_OPPONENT_TYPES = {'Clearance', 'BlockedPass', 'Save'}
# Same-team shot events that can turn up as the action immediately before
# ANOTHER shot - a rebound. In that case the earlier shot IS the SCA (labeled
# simply "Shot", not the specific outcome type), and the search stops right
# there - a rebound never gets a second SCA behind it.
SHOT_TYPES = {'SavedShot', 'MissedShots', 'ShotOnPost', 'Goal'}

# "Open play" = not a corner, free kick (direct/indirect), throw-in, goal kick,
# or keeper throw. Used to restrict Progressive Passes and the Passing tab's
# open-play-only columns to the same definition.
OPEN_PLAY_EXCLUDE_QUALIFIERS = {'CornerTaken', 'FreekickTaken', 'IndirectFreekickTaken',
                                 'ThrowIn', 'GoalKick', 'KeeperThrow'}


# ============================================================
# 0. FIXTURE DISCOVERY (batch runner support)
# ============================================================
# Matches WhoScored's plain-text date headers on a fixtures/results page,
# e.g. "Saturday, Aug 22 2026".
_FIXTURE_DATE_HEADER_RE = re.compile(
    r"^(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),\s+\w+\s+\d{1,2}\s+\d{4}$"
)
# Status tokens WhoScored shows next to a FINISHED match - confirmed by
# fetching a real, already-played month of Premier League fixtures (every
# match on that page showed "FT" as its own standalone token; matches that
# went to extra time/penalties are expected to use "AET"/"Pens" the same
# way, though that specific case hasn't been seen in real data yet). Any
# other token next to a match (a kickoff time, "--", "LIVE", a live match
# minute, ...) is treated as NOT finished - the safe default for a batch
# runner that should only auto-run matches that have actually been played.
_FIXTURE_FINISHED_STATUS_WORDS = {"FT", "AET", "PENS", "PEN"}


def get_fixture_urls(fixtures_url, only_finished=True):
    """
    Scraper for a WhoScored competition fixtures/results page (e.g. the
    Premier League's own "Fixtures" tab) - returns a list of {match_url,
    home_name, away_name, status, date} dicts, one per match found on that
    page. Built to support a "run the whole weekend's matches" batch
    workflow, so you don't have to copy/paste each match URL by hand.

    REAL PAGE STRUCTURE THIS IS BUILT AGAINST: confirmed by fetching two
    real pages - an upcoming/not-yet-played fixtures page, and (separately)
    a real month of ALREADY-PLAYED Premier League results. Neither has any
    embedded JSON like the match-centre page scrape_match() reads above -
    fixtures are plain server-rendered HTML. Each match has a short status
    token as standalone text immediately before it ("FT" for a finished
    match - confirmed directly from the real results page; something else -
    a kickoff time, "--", etc. - for a not-yet-played one), followed by a
    link to the match itself and then two
    <a href="/teams/<id>/show/<team-slug>"> links giving the home then away
    team name as their visible text. Date headers appear as standalone text
    ("Saturday, May 02 2026") preceding each day's block of matches. This
    function walks the parsed HTML in document order, tracking the most
    recent date header and status token, and pairing each match link with
    the two team links that follow it.

    On the real already-played results page, a finished match's own link
    already points straight at the "/live/" match-centre URL
    scrape_match() needs - this used to be an unconfirmed guess (the
    fixture list was assumed to only expose a "/show/" preview-page link),
    but real data now confirms it directly. A not-yet-played match's link
    still uses "/show/", which this function rewrites to the equivalent
    "/live/" URL for consistency - though with only_finished=True (the
    default) that rewritten URL is normally filtered out before you'd ever
    use it, since it points at a match that hasn't happened yet.

    only_finished: if True (default), keeps only matches whose status token
    is a recognized "finished" marker (see _FIXTURE_FINISHED_STATUS_WORDS).
    """
    print(f"Opening {fixtures_url} ...")
    with get_driver() as driver:
        driver.get(fixtures_url)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        page_source = driver.page_source

    soup = BeautifulSoup(page_source, "html.parser")

    all_matches = []
    seen_urls = set()
    current_date = None
    pending_status = None  # most recent short status token seen, consumed by the next match link
    pending = None  # the match dict currently waiting on its two team-name links

    for el in soup.descendants:
        if isinstance(el, NavigableString):
            text = str(el).strip()
            if not text:
                continue
            if _FIXTURE_DATE_HEADER_RE.match(text):
                current_date = text
                continue
            # A short, single-token line (no spaces) is this match's status
            # ("FT", a kickoff time, "--", ...) - but only capture it while
            # we're between matches. In between a match's own two team links,
            # WhoScored also shows the team's score as a bare digit
            # ("Brentford" / "1" / "West Ham") - ignoring stray text while
            # `pending` is set keeps a score digit like that from being
            # mistaken for the NEXT match's status.
            if pending is None and " " not in text and len(text) <= 12:
                pending_status = text
            continue

        if getattr(el, "name", None) != "a":
            continue
        href = (el.get("href") or "").split("#", 1)[0].rstrip("/")

        m = re.match(r"^/matches/(\d+)/(?:live|show)/([^/]+)$", href)
        if m:
            match_id, slug = m.groups()
            match_url = f"https://www.whoscored.com/matches/{match_id}/live/{slug}"
            if match_url in seen_urls:
                # A later "comments"/"stats" link back to a match already
                # recorded (both the /show/ and /live/ variants normalize to
                # the same match_url above) - not a new match.
                continue
            seen_urls.add(match_url)
            pending = {
                "match_url": match_url,
                "status": pending_status,
                "date": current_date,
                "home_name": None,
                "away_name": None,
            }
            all_matches.append(pending)
            pending_status = None
            continue

        if pending is not None and re.match(r"^/teams/\d+/show/", href):
            if pending["home_name"] is None:
                pending["home_name"] = el.get_text(strip=True)
            elif pending["away_name"] is None:
                pending["away_name"] = el.get_text(strip=True)
                pending = None  # both team names found, stop touching this match

    if not all_matches:
        raise RuntimeError(
            "Couldn't find any fixture links on this page (looked for "
            "/matches/<id>/live/... or /show/... anchors). WhoScored may have "
            "changed this page's structure since get_fixture_urls() was last "
            "checked against it, or this URL doesn't show a fixture list at "
            "all. Paste match URLs manually instead of using auto-detect for "
            "now, and share the page's HTML so this can be fixed against real "
            "data."
        )

    results = []
    for m in all_matches:
        if not (m["home_name"] and m["away_name"]):
            # Couldn't confidently pair this match link with two team names -
            # skip rather than guess at who played.
            continue
        status_word = (m["status"] or "").upper()
        is_finished = status_word in _FIXTURE_FINISHED_STATUS_WORDS
        if only_finished and not is_finished:
            continue
        results.append({
            "match_url": m["match_url"],
            "home_name": m["home_name"],
            "away_name": m["away_name"],
            "status": m["status"] or "(unknown)",
            "date": m["date"],
        })

    return results


# ============================================================
# 1. SCRAPE
# ============================================================
def scrape_match(match_centre_url):
    """
    Load a WhoScored match centre page and pull out both the event stream
    and the home/away team names, using the same Selenium approach as
    whoscored/whoscored_events_data.py in this repo, extended to also
    read team names directly from the match JSON (rather than guessing).
    """
    print(f"Opening {match_centre_url} ...")
    with get_driver() as driver:
        driver.get(match_centre_url)
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        page_source = driver.page_source

    soup = BeautifulSoup(page_source, "html.parser")
    script_tag = soup.select_one('script:-soup-contains("matchCentreData")')
    if not script_tag:
        raise RuntimeError(
            "No script tag with matchCentreData found - WhoScored may have "
            "blocked the automated browser or changed its page structure."
        )

    _, _, json_text = script_tag.text.partition("matchCentreData: ")
    match_json = json.loads(json_text.split(",\n")[0])

    player_id_name_dict = match_json.get("playerIdNameDictionary", {})
    events_dict = match_json.get("events", {})
    if not events_dict:
        raise RuntimeError("No events found in matchCentreData for this match.")

    df = pd.json_normalize(events_dict)

    if "playerId" in df.columns:
        df["playerName"] = df["playerId"].apply(
            lambda x: player_id_name_dict.get(str(int(x))) if pd.notna(x) else None
        )
    if "relatedPlayerId" in df.columns:
        df["relatedPlayerName"] = df["relatedPlayerId"].apply(
            lambda x: player_id_name_dict.get(str(int(x))) if pd.notna(x) else None
        )

    # Team names, read directly from the match JSON rather than guessed.
    home = match_json.get("home", {})
    away = match_json.get("away", {})
    team_map = {}
    home_name, away_name = None, None
    if home.get("teamId") is not None and home.get("name"):
        team_map[home["teamId"]] = home["name"]
        home_name = home["name"]
    if away.get("teamId") is not None and away.get("name"):
        team_map[away["teamId"]] = away["name"]
        away_name = away["name"]

    if not team_map:
        print("WARNING: could not read team names from the page - falling "
              "back to generic 'Team <id>' labels. Check scrape_match() "
              "if this happens.")
        ids = sorted(df['teamId'].dropna().unique())
        team_map = {tid: f"Team {int(tid)}" for tid in ids}
        if len(ids) >= 2:
            home_name, away_name = team_map[ids[0]], team_map[ids[1]]

    df['team'] = df['teamId'].map(team_map)
    df = df.reset_index(drop=True)
    df['qualifiers_parsed'] = df['qualifiers'].apply(
        lambda q: q if isinstance(q, list) else (ast.literal_eval(q) if isinstance(q, str) else [])
    )

    match_info = {'home_name': home_name, 'away_name': away_name}
    return df, match_info


# ============================================================
# 2. GEOMETRY HELPERS
# ============================================================
def to_yd_x(x):
    return x / 100.0 * PITCH_LEN_M * M_TO_YD


def to_yd_y(y):
    return y / 100.0 * PITCH_WID_M * M_TO_YD


GOAL_X_YD = to_yd_x(100)
GOAL_Y_YD = to_yd_y(50)
BOX_X_MIN = GOAL_X_YD - BOX_DEPTH_YD
BOX_Y_MIN = GOAL_Y_YD - BOX_WIDTH_YD / 2
BOX_Y_MAX = GOAL_Y_YD + BOX_WIDTH_YD / 2


def dist_to_goal_yd(x, y):
    return np.sqrt((GOAL_X_YD - to_yd_x(x)) ** 2 + (GOAL_Y_YD - to_yd_y(y)) ** 2)


def in_box(x, y):
    return (to_yd_x(x) >= BOX_X_MIN) & (to_yd_y(y) >= BOX_Y_MIN) & (to_yd_y(y) <= BOX_Y_MAX)


def qual_names(qlist):
    return {item['type']['displayName'] for item in qlist}


def third(x):
    if x < OWN_THIRD_MAX:
        return 'Own third'
    elif x < MIDDLE_THIRD_MAX:
        return 'Middle third'
    else:
        return 'Final third'


# ============================================================
# 3. PROGRESSIVE PASSES
# ============================================================
def _pass_receiver_map(df):
    """
    WhoScored/Opta pass events carry no explicit "receiver" field, so the
    receiver is inferred the same way most open-source pass-network scripts
    do it: the player of the very next event in the full chronological log,
    provided that next event belongs to the SAME team (otherwise the pass
    wasn't genuinely received by that team, so no receiver is recorded).
    Returns a Series of receiver names, aligned to df's own index.
    """
    work = df.sort_index()
    receiver = work['playerName'].shift(-1)
    receiver_team = work['team'].shift(-1)
    valid = receiver_team == work['team']
    return receiver.where(valid)


def _compute_progressive_flags(df):
    """
    Shared core of the progressive-pass rolling-window algorithm, factored
    out of compute_progressive_passes() so other consumers (e.g. the Pass
    Map in streamlit_app.py) can mark which individual passes are
    progressive without re-deriving the rolling baseline differently and
    risking drift between the two. Returns one row per completed pass with
    valid end coordinates (boundary events dropped), including an
    'orig_index' column pointing back at the row's position in the original
    df, so callers can look a specific event back up by df.index.

    Only OPEN PLAY completed passes can be flagged as progressive - corners,
    free kicks (direct/indirect), throw-ins, goal kicks, and keeper throws
    are excluded from the count, even if they'd otherwise meet the distance/
    box criteria. They're still fed into the rolling 6-pass baseline used to
    measure "progress" (an attacking team's furthest reach doesn't reset
    just because the delivery was a set piece), but they can never
    themselves be counted as a progressive pass.
    """
    receiver_map = _pass_receiver_map(df)
    work = df[~df['type.displayName'].isin(BOUNDARY_TYPES)].copy()
    work = work.sort_index()

    rolling = {t: deque(maxlen=PROGRESSIVE_ROLLING_WINDOW) for t in work['team'].dropna().unique()}
    prog_pass_records = []

    for i in range(len(work)):
        row = work.iloc[i]
        team = row['team']
        if pd.isna(team):
            continue
        if (row['type.displayName'] == 'Pass' and row['outcomeType.displayName'] == 'Successful'
                and pd.notna(row['endX']) and pd.notna(row['endY'])):
            start_dist = dist_to_goal_yd(row['x'], row['y'])
            end_dist = dist_to_goal_yd(row['endX'], row['endY'])
            win = rolling[team]
            baseline = min(win) if len(win) > 0 else start_dist
            progress = baseline - end_dist
            ends_in_box = in_box(row['endX'], row['endY'])
            starts_def = row['x'] < DEFENSIVE_START_CUTOFF
            is_open_play = not bool(OPEN_PLAY_EXCLUDE_QUALIFIERS & qual_names(row['qualifiers_parsed']))
            is_progressive = is_open_play and \
                              ((progress >= PROGRESSIVE_YARD_THRESHOLD) or ends_in_box) and \
                              ((not starts_def) or ends_in_box)
            prog_pass_records.append({
                'orig_index': row.name,
                'minute': row['minute'], 'second': row['second'], 'team': team,
                'player': row['playerName'], 'receiver': receiver_map.get(row.name),
                'x': row['x'], 'y': row['y'],
                'endX': row['endX'], 'endY': row['endY'],
                'progress_yd': round(progress, 1), 'ends_in_box': bool(ends_in_box),
                'is_progressive': bool(is_progressive),
            })
            win.append(end_dist)

    return pd.DataFrame(prog_pass_records)


def compute_progressive_passes(df):
    """
    Only OPEN PLAY completed passes can be flagged as progressive - see
    _compute_progressive_flags() above for the full rolling-window
    definition/logic this reuses.
    """
    prog_passes_all = _compute_progressive_flags(df)
    prog_passes = prog_passes_all[prog_passes_all['is_progressive']].copy()
    prog_passes_out = prog_passes[
        ['minute', 'second', 'team', 'player', 'x', 'y', 'endX', 'endY', 'progress_yd', 'ends_in_box']
    ]

    player_totals = (prog_passes.groupby(['team', 'player']).size()
                      .reset_index(name='progressive_passes')
                      .sort_values('progressive_passes', ascending=False))
    team_totals = (prog_passes.groupby('team').size()
                   .reset_index(name='progressive_passes')
                   .sort_values('progressive_passes', ascending=False))
    player_received = (prog_passes.dropna(subset=['receiver']).groupby(['team', 'receiver']).size()
                        .reset_index(name='progressive_passes_received')
                        .rename(columns={'receiver': 'player'}))

    return prog_passes_out, player_totals, team_totals, player_received


def compute_passes_received(df):
    """
    Passes Received per player: completed passes where this player was the
    receiver, using the same next-event heuristic as _pass_receiver_map()
    (WhoScored/Opta has no explicit receiver field on pass events).
    """
    receiver_map = _pass_receiver_map(df)
    work = df.sort_index().copy()
    work['receiver'] = receiver_map
    pass_mask = (work['type.displayName'] == 'Pass') & (work['outcomeType.displayName'] == 'Successful')
    received = work[pass_mask].dropna(subset=['receiver'])
    return (received.groupby(['team', 'receiver']).size()
            .reset_index(name='passes_received')
            .rename(columns={'receiver': 'player'}))


def compute_passing_pairs(df):
    """
    Passing Pairs tab: every distinct passer -> receiver combination
    (completed passes only, any pass type - open play, corners, free kicks,
    etc.), with a count of how many times that exact pair happened, split
    by team. Receiver uses the same next-event, same-team heuristic as
    _pass_receiver_map()/compute_passes_received() - WhoScored/Opta pass
    events carry no explicit receiver field. Sorted within each team by
    count, descending, so the most frequent combinations come first.
    """
    receiver_map = _pass_receiver_map(df)
    work = df.sort_index().copy()
    work['receiver'] = receiver_map
    pass_mask = (work['type.displayName'] == 'Pass') & (work['outcomeType.displayName'] == 'Successful')
    completed = work[pass_mask].dropna(subset=['receiver'])
    pairs = (completed.groupby(['team', 'playerName', 'receiver']).size()
             .reset_index(name='count')
             .rename(columns={'playerName': 'passer'}))
    pairs = pairs.sort_values(['team', 'count'], ascending=[True, False]).reset_index(drop=True)
    return pairs[['team', 'passer', 'receiver', 'count']]


# ============================================================
# 3a. PASS MAP (single-player pass events, for the Streamlit pass map)
# ============================================================
def classify_pass_category(row):
    """
    One of four mutually-exclusive pass-map categories, in priority order:
    an unsuccessful pass is always 'Incomplete' regardless of anything else
    (key pass and progressive both require a completed pass by definition,
    so they never apply to an incomplete one). A completed pass that
    qualifies as BOTH a key pass (ShotAssist) and progressive is labeled
    'Key Pass' - the rarer, more specific category wins over
    the broader 'Progressive' one. Everything else completed is 'Completed'.
    """
    if not row['completed']:
        return 'Incomplete'
    if row['is_key_pass']:
        return 'Key Pass'
    if row['is_progressive']:
        return 'Progressive'
    return 'Completed'


def get_player_passes(df, player):
    """
    Every Pass event (any outcome, open play or set piece) attempted by one
    player, with the flags needed to color a pass map: completed (successful
    outcome), is_progressive (the exact same rolling-window definition/logic
    used for the Progressive Passes column elsewhere in this workbook, via
    _compute_progressive_flags() - note
    this is always False for an incomplete or set-piece pass, since the
    Progressive Passes definition only ever applies to completed, open-play
    passes), and is_key_pass (WhoScored's own 'ShotAssist' qualifier - same
    source as the Shot Assists column on the Passing tab). 'category' is the
    single mutually-exclusive label from classify_pass_category() above,
    meant for coloring a pass map.
    """
    out_cols = ['minute', 'second', 'team', 'player', 'x', 'y', 'endX', 'endY',
                'completed', 'is_progressive', 'is_key_pass', 'category']

    work = df[(df['type.displayName'] == 'Pass') & (df['playerName'] == player)].copy()
    if work.empty:
        return pd.DataFrame(columns=out_cols)

    prog_flags = _compute_progressive_flags(df)
    prog_lookup = (prog_flags.set_index('orig_index')['is_progressive']
                   if not prog_flags.empty else pd.Series(dtype=bool))

    work['completed'] = work['outcomeType.displayName'] == 'Successful'
    work['is_key_pass'] = work['qualifiers_parsed'].apply(lambda qs: 'ShotAssist' in qual_names(qs))
    work['is_progressive'] = [bool(prog_lookup.get(i, False)) for i in work.index]
    work['category'] = work.apply(classify_pass_category, axis=1)
    work = work.rename(columns={'playerName': 'player'})

    return work[out_cols].sort_values(['minute', 'second']).reset_index(drop=True)


def get_player_passes_received(df, player):
    """
    Every COMPLETED Pass received by one player - the mirror image of
    get_player_passes() above, for the Passes Received map in
    streamlit_app.py. Receiver is inferred the same way as everywhere else
    in this module (_pass_receiver_map() - the next event's player, same
    team), restricted up front to Successful passes only, since an
    incomplete pass was never actually received by anyone on the attacking
    team (there's no "receiver" to speak of) - so unlike get_player_passes(),
    'category' here is never 'Incomplete'.

    x/y are the PASSER's origin (where the ball came from); endX/endY are
    where this player actually received it - so a map can still draw a line
    from origin to reception point, same as the outgoing Pass Map, just
    from the receiving player's perspective instead of the passer's.
    is_progressive/is_key_pass/category use the exact same definitions as
    get_player_passes() (a pass's progressive/key-pass status doesn't
    depend on which end of it you're looking from).
    """
    out_cols = ['minute', 'second', 'team', 'passer', 'player', 'x', 'y', 'endX', 'endY',
                'completed', 'is_progressive', 'is_key_pass', 'category']

    receiver_map = _pass_receiver_map(df)
    work = df[(df['type.displayName'] == 'Pass') & (df['outcomeType.displayName'] == 'Successful')].copy()
    work['receiver'] = receiver_map.reindex(work.index)
    work = work[work['receiver'] == player]
    if work.empty:
        return pd.DataFrame(columns=out_cols)

    prog_flags = _compute_progressive_flags(df)
    prog_lookup = (prog_flags.set_index('orig_index')['is_progressive']
                   if not prog_flags.empty else pd.Series(dtype=bool))

    work['completed'] = True  # already restricted to Successful above
    work['is_key_pass'] = work['qualifiers_parsed'].apply(lambda qs: 'ShotAssist' in qual_names(qs))
    work['is_progressive'] = [bool(prog_lookup.get(i, False)) for i in work.index]
    work['category'] = work.apply(classify_pass_category, axis=1)
    work = work.rename(columns={'playerName': 'passer', 'receiver': 'player'})

    return work[out_cols].sort_values(['minute', 'second']).reset_index(drop=True)


def compute_all_passes(df):
    """
    Every Pass event in the match (any player, any outcome), with the same
    completed/is_progressive/is_key_pass/category flags as get_player_passes()
    and get_player_passes_received() - the shared source both of those are
    really just a player-filtered view of. Computed once for the whole match
    (not per-player) so publishing every player's passes to the history
    database doesn't mean re-running the progressive-passes rolling window
    once per player.
    """
    out_cols = ['minute', 'second', 'team', 'passer', 'receiver', 'x', 'y', 'endX', 'endY',
                'completed', 'is_progressive', 'is_key_pass', 'category']
    work = df[df['type.displayName'] == 'Pass'].copy()
    if work.empty:
        return pd.DataFrame(columns=out_cols)

    receiver_map = _pass_receiver_map(df)
    prog_flags = _compute_progressive_flags(df)
    prog_lookup = (prog_flags.set_index('orig_index')['is_progressive']
                   if not prog_flags.empty else pd.Series(dtype=bool))

    work['completed'] = work['outcomeType.displayName'] == 'Successful'
    work['is_key_pass'] = work['qualifiers_parsed'].apply(lambda qs: 'ShotAssist' in qual_names(qs))
    work['is_progressive'] = [bool(prog_lookup.get(i, False)) for i in work.index]
    work['category'] = work.apply(classify_pass_category, axis=1)
    work['receiver'] = receiver_map.reindex(work.index)
    # An incomplete pass's "receiver" (same-team next event) isn't really a
    # receiver in any meaningful sense - null it out so downstream Passes
    # Received queries (which should only ever show completed passes) can't
    # accidentally pick one up.
    work.loc[~work['completed'], 'receiver'] = None
    work = work.rename(columns={'playerName': 'passer'})

    return work[out_cols].sort_values(['minute', 'second']).reset_index(drop=True)


# ============================================================
# 3b. CARRIES
# ============================================================
PERIOD_ORDER = {'FirstHalf': 1, 'SecondHalf': 2, 'FirstPeriodOfExtraTime': 3,
                'SecondPeriodOfExtraTime': 4, 'PenaltyShootout': 5, 'PostGame': 14, 'PreMatch': 16}

CARRY_MIN_LEN = 3.0    # metres
CARRY_MAX_LEN = 60.0   # metres
CARRY_MIN_DUR = 1.0    # seconds
CARRY_MAX_DUR = 10.0   # seconds
CARRY_SKIP_TYPES = {'TakeOn', 'Foul'}  # events to look past when searching for a carry's endpoint


def _add_cumulative_mins(df):
    """Elapsed match time in minutes, running continuously across periods
    (so second-half minute 46 continues on from first-half minute 45, etc.)."""
    d = df.copy()
    d['period_num'] = d['period.displayName'].map(PERIOD_ORDER)
    d['cumulative_mins'] = d['minute'] + d['second'] / 60.0
    periods_present = sorted(d['period_num'].dropna().unique())
    for i, period in enumerate(periods_present):
        if i == 0:
            continue
        prev_period = periods_present[i - 1]
        prev_max = d.loc[d['period_num'] == prev_period, 'cumulative_mins'].max()
        curr_min = d.loc[d['period_num'] == period, 'cumulative_mins'].min()
        d.loc[d['period_num'] == period, 'cumulative_mins'] += (prev_max - curr_min)
    return d


def compute_carries(df):
    """
    WhoScored/Opta data has no explicit "carry" event, so carries are
    inferred from gaps between a team's consecutive logged actions - this
    follows the same approach used by the well-known open-source WhoScored
    match-report notebooks (e.g. jakeyk11 / adnaaan433's Post-Match-Report),
    so that results line up with other tools built on the same methodology:
    for each event, look ahead to the next event, skipping past TakeOns and
    Fouls (which don't represent a genuine loss of the ball), then treat the
    gap between the first event's end location and that next event's start
    location as a carry, provided: both belong to the same team, the first
    event isn't itself a bare "BallTouch", the distance covered is between
    3 and 60 metres, and the elapsed time is between 1 and 10 seconds (to
    rule out carries that are too trivial, too implausibly long, or that
    silently bridge a stoppage).

    Returns (team_carries, player_carries, carries_df): the first two are
    whole-match totals (per team / per team+player); carries_df is the raw,
    one-row-per-carry data behind them (team, player, cumulative_mins,
    is_progressive, into_final_third, into_box) - kept around so
    compute_on_off() can window carries down to a specific player's own
    on-pitch time, which the two aggregate tables alone can't do.
    """
    d = _add_cumulative_mins(df)
    n = len(d)
    carry_records = []

    for idx in range(n - 1):
        cur = d.iloc[idx]
        cur_team = cur['teamId']

        j = idx + 1
        nxt = None
        while j < n:
            cand = d.iloc[j]
            if cand['type.displayName'] == 'TakeOn' and cand['outcomeType.displayName'] == 'Successful':
                j += 1
                continue
            if ((cand['type.displayName'] == 'TakeOn' and cand['outcomeType.displayName'] == 'Unsuccessful')
                    or (cand['teamId'] != cur_team and cand['type.displayName'] == 'Challenge'
                        and cand['outcomeType.displayName'] == 'Unsuccessful')
                    or (cand['type.displayName'] == 'Foul')):
                j += 1
                continue
            nxt = cand
            break
        if nxt is None:
            continue

        same_team = cur_team == nxt['teamId']
        not_ball_touch = cur['type.displayName'] != 'BallTouch'
        if pd.isna(cur.get('endX')) or pd.isna(cur.get('endY')):
            continue
        dx = 105 * (cur['endX'] - nxt['x']) / 100
        dy = 68 * (cur['endY'] - nxt['y']) / 100
        far_enough = dx ** 2 + dy ** 2 >= CARRY_MIN_LEN ** 2
        not_too_far = dx ** 2 + dy ** 2 <= CARRY_MAX_LEN ** 2
        dt = 60 * (nxt['cumulative_mins'] - cur['cumulative_mins'])
        min_time = dt >= CARRY_MIN_DUR
        same_phase = dt < CARRY_MAX_DUR
        same_period = cur['period.displayName'] == nxt['period.displayName']

        if not (same_team and not_ball_touch and far_enough and not_too_far
                and min_time and same_phase and same_period):
            continue

        sx, sy = cur['endX'], cur['endY']
        ex, ey = nxt['x'], nxt['y']
        # Distance-to-goal comparison uses the same 105x68 pitch scaling as
        # the carry-length check above (rather than yards) to stay
        # consistent with this specific methodology.
        start_dist = np.sqrt((105 - sx * 1.05) ** 2 + (34 - sy * 0.68) ** 2)
        end_dist = np.sqrt((105 - ex * 1.05) ** 2 + (34 - ey * 0.68) ** 2)
        progress = start_dist - end_dist
        ends_in_box = in_box(ex, ey)
        starts_in_box = in_box(sx, sy)
        start_third = third(sx)
        end_third = third(ex)

        # Progressive: moves the ball >=10 (of these units) closer to goal,
        # or enters the box. Carries starting in a team's own third still
        # count, but carries ENDING in their own third are excluded.
        is_progressive = ((progress >= PROGRESSIVE_YARD_THRESHOLD) or ends_in_box) and \
                          (end_third != 'Own third')

        carry_records.append({
            'team': cur['team'], 'player': nxt['playerName'],
            'cumulative_mins': cur['cumulative_mins'],
            'is_progressive': bool(is_progressive),
            'into_final_third': bool(end_third == 'Final third' and start_third != 'Final third'),
            'into_box': bool(ends_in_box and not starts_in_box),
        })

    carries_df = pd.DataFrame(carry_records) if carry_records else pd.DataFrame(
        columns=['team', 'player', 'cumulative_mins', 'is_progressive', 'into_final_third', 'into_box'])

    team_carries = carries_df.groupby('team').agg(
        progressive_carries=('is_progressive', 'sum'),
        carries_into_final_third=('into_final_third', 'sum'),
        carries_into_box=('into_box', 'sum'),
    ).reset_index() if len(carries_df) else pd.DataFrame(
        columns=['team', 'progressive_carries', 'carries_into_final_third', 'carries_into_box'])

    player_carries = carries_df.groupby(['team', 'player']).agg(
        progressive_carries=('is_progressive', 'sum'),
        carries_into_final_third=('into_final_third', 'sum'),
        carries_into_box=('into_box', 'sum'),
    ).reset_index() if len(carries_df) else pd.DataFrame(
        columns=['team', 'player', 'progressive_carries', 'carries_into_final_third', 'carries_into_box'])

    # carries_df (the raw, one-row-per-carry data, with each carry's own
    # cumulative_mins) is also returned so compute_on_off() can window
    # carries by a specific player's on-pitch time - team_carries/
    # player_carries above only have whole-match totals, which isn't enough
    # for an on/off split.
    return team_carries, player_carries, carries_df


# ============================================================
# 4. SHOT-CREATING ACTIONS (+ shot distance, body part)
# ============================================================
def classify_action(row):
    rtype = row['type.displayName']
    if rtype == 'TakeOn':
        return 'take-on'
    if rtype in SHOT_TYPES:
        return 'Shot'
    if rtype != 'Pass':
        return rtype
    qn = qual_names(row['qualifiers_parsed'])
    if 'CornerTaken' in qn:
        return 'Pass (Corner)'
    if 'FreekickTaken' in qn or 'IndirectFreekickTaken' in qn:
        return 'Pass (Set Piece)'
    if 'ThrowIn' in qn:
        return 'Pass (Throw-in)'
    if 'GoalKick' in qn:
        return 'Pass (Goal Kick)'
    if 'KeeperThrow' in qn:
        return 'Pass (Keeper Throw)'
    return 'Pass (Live)'


def body_part(row):
    qn = qual_names(row['qualifiers_parsed'])
    if 'Head' in qn:
        return 'Head'
    if 'LeftFoot' in qn:
        return 'Left Foot'
    if 'RightFoot' in qn:
        return 'Right Foot'
    return 'Other'


def is_own_goal(row):
    return row['type.displayName'] == 'Goal' and 'Own goal' in qual_names(row['qualifiers_parsed'])


def compute_sca(df):
    # Own goals carry WhoScored's 'isShot' flag like any other Goal event,
    # but they aren't a shot taken by the credited team - exclude them so
    # they don't count as a shot or get their own SCA row.
    own_goal_mask = df.apply(is_own_goal, axis=1)
    shots_idx = df.index[(df['isShot'] == True) & ~own_goal_mask].tolist()
    sca_rows = []
    for shot_i in shots_idx:
        shot_team = df.at[shot_i, 'team']
        found = []
        j = shot_i - 1
        while j >= 0 and len(found) < 2:
            row = df.loc[j]
            rtype = row['type.displayName']
            rteam = row['team']
            routcome = row['outcomeType.displayName']

            if rtype in BOUNDARY_TYPES:
                break

            if rtype in NON_BREAKING_DUEL_TYPES:
                if rtype in COUNTABLE_DUEL_TYPES and rteam == shot_team and routcome == 'Successful':
                    found.append(j)
                j -= 1
                continue

            if rteam == shot_team:
                if rtype in OFFENSIVE_TYPES:
                    found.append(j)
                    if rtype in SHOT_TYPES:
                        # This shot is a rebound off the previous one - that
                        # previous shot IS the SCA, and a rebound never gets
                        # a second SCA behind it, so stop here.
                        break
            else:
                if rtype == 'Foul':
                    found.append(j)
                    break
                elif rtype in LOOSE_BALL_OPPONENT_TYPES:
                    # A partial clearance/block/save can still put the ball
                    # right back to the shooting team - keep searching past
                    # it for the real originating pass, without counting the
                    # defensive touch itself as an SCA.
                    j -= 1
                    continue
                else:
                    break
            j -= 1

        shot_row = df.loc[shot_i]
        rec = {
            'minute': shot_row['minute'],
            'team': shot_team,
            'player': shot_row['playerName'],
            'shotType': shot_row['type.displayName'],
            'shot_distance_yd': round(dist_to_goal_yd(shot_row['x'], shot_row['y']), 1),
            'bodyPart': body_part(shot_row),
            'sca1_player': None, 'sca1_action': None,
            'sca2_player': None, 'sca2_action': None,
        }
        for rank, action_i in enumerate(found, start=1):
            arow = df.loc[action_i]
            rec[f'sca{rank}_player'] = arow['playerName']
            rec[f'sca{rank}_action'] = classify_action(arow)
        sca_rows.append(rec)

    return pd.DataFrame(sca_rows)


def compute_shot_pairs(sca_out):
    """
    Shot Pairs tab: every distinct passer -> shot-taker combination, with a
    count of how many times that exact pair happened, split by team. The
    passer is whoever played the SCA1 action (from compute_sca() above) for
    that shot, and only counts when that immediate preceding action was
    itself a pass (any type - open play, corner, free kick, etc., i.e. any
    classify_action() value starting with "Pass"). Shots whose SCA1 was a
    take-on, duel, rebound off a blocked/saved shot, or a loose ball with no
    such preceding pass aren't a passer -> shooter "pair" and are excluded.
    Sorted within each team by count, descending, so the most frequent
    combinations come first.
    """
    work = sca_out.copy()
    pass_mask = work['sca1_action'].astype(str).str.startswith('Pass')
    paired = work[pass_mask].dropna(subset=['sca1_player'])
    pairs = (paired.groupby(['team', 'player', 'sca1_player']).size()
             .reset_index(name='count')
             .rename(columns={'player': 'shot_taker', 'sca1_player': 'passer'}))
    pairs = pairs.sort_values(['team', 'count'], ascending=[True, False]).reset_index(drop=True)
    return pairs[['team', 'shot_taker', 'passer', 'count']]


# ============================================================
# 5. TOUCHES (renamed from "Touches by Third"), with carries merged in
# ============================================================
def compute_touches(df, team_carries, player_carries, passes_received, progressive_received):
    touches = df[df['isTouch'] == True].copy()
    touches['pitch_third'] = touches['x'].apply(third)
    touches['in_att_box'] = in_box(touches['x'], touches['y'])

    team_summary = touches.groupby(['team', 'pitch_third']).size().unstack(fill_value=0)
    for c in ['Own third', 'Middle third', 'Final third']:
        if c not in team_summary.columns:
            team_summary[c] = 0
    team_summary = team_summary[['Own third', 'Middle third', 'Final third']]
    team_summary['Attacking Box'] = touches[touches['in_att_box']].groupby('team').size()
    team_summary['Attacking Box'] = team_summary['Attacking Box'].fillna(0).astype(int)
    team_summary['Total touches'] = team_summary[['Own third', 'Middle third', 'Final third']].sum(axis=1)
    team_summary = team_summary.reset_index().merge(team_carries, on='team', how='left')
    for c in ['progressive_carries', 'carries_into_final_third', 'carries_into_box']:
        team_summary[c] = team_summary[c].fillna(0).astype(int)
    team_summary = team_summary.rename(columns={
        'progressive_carries': 'Progressive Carries',
        'carries_into_final_third': 'Carries into Final Third',
        'carries_into_box': 'Carries into Box',
    })
    for c in ['Own third', 'Middle third', 'Final third', 'Attacking Box']:
        team_summary[c + ' %'] = (team_summary[c] / team_summary['Total touches'] * 100).round(1)
    team_summary = team_summary[[
        'team', 'Total touches', 'Own third', 'Middle third', 'Final third', 'Attacking Box',
        'Progressive Carries', 'Carries into Final Third', 'Carries into Box',
        'Own third %', 'Middle third %', 'Final third %', 'Attacking Box %',
    ]]

    player_third = touches.groupby(['team', 'playerName', 'pitch_third']).size().unstack(fill_value=0)
    for c in ['Own third', 'Middle third', 'Final third']:
        if c not in player_third.columns:
            player_third[c] = 0
    player_third = player_third[['Own third', 'Middle third', 'Final third']]
    att_box_by_player = touches[touches['in_att_box']].groupby(['team', 'playerName']).size()
    player_third['Attacking Box'] = att_box_by_player
    player_third['Attacking Box'] = player_third['Attacking Box'].fillna(0).astype(int)
    player_third['Total Touches'] = player_third[['Own third', 'Middle third', 'Final third']].sum(axis=1)
    player_third = player_third.reset_index().rename(columns={'playerName': 'player'})
    player_third = player_third.merge(player_carries, on=['team', 'player'], how='left')
    for c in ['progressive_carries', 'carries_into_final_third', 'carries_into_box']:
        player_third[c] = player_third[c].fillna(0).astype(int)
    player_third = player_third.rename(columns={
        'progressive_carries': 'Progressive Carries',
        'carries_into_final_third': 'Carries into Final Third',
        'carries_into_box': 'Carries into Box',
    })
    player_third = player_third.merge(passes_received, on=['team', 'player'], how='left')
    player_third = player_third.merge(progressive_received, on=['team', 'player'], how='left')
    for c in ['passes_received', 'progressive_passes_received']:
        player_third[c] = player_third[c].fillna(0).astype(int)
    player_third = player_third.rename(columns={
        'passes_received': 'Passes Received',
        'progressive_passes_received': 'Progressive Passes Received',
    })

    player_third = player_third[[
        'team', 'player', 'Total Touches', 'Own third', 'Middle third', 'Final third', 'Attacking Box',
        'Progressive Carries', 'Carries into Final Third', 'Carries into Box',
        'Passes Received', 'Progressive Passes Received',
    ]]
    player_third = player_third.sort_values(['team', 'Total Touches'], ascending=[True, False])

    return team_summary, player_third


# ============================================================
# 5b. PASSING
# ============================================================
def _classify_passes(df):
    """
    Shared pass-event classification, used by both compute_passing() (per-
    player pass totals) and compute_on_off() (windowed On/Off splits) - one
    canonical definition of "open play cross" / "into final third" / "into
    the box" for a pass, so those two tabs can never quietly disagree with
    each other about what counts. 'Open play' excludes corners and free
    kicks (direct or indirect) - see OPEN_PLAY_EXCLUDE_QUALIFIERS above.

    Returns just the Pass rows from df (with cumulative_mins already
    attached via _add_cumulative_mins, since compute_on_off needs each
    pass's own match minute to know which player's on-pitch window it
    falls in), plus these added columns: qn (qualifier name set),
    completed, forward, headed, is_cross, is_open_play, open_play_cross,
    open_play_cross_completed, into_final_third, into_box, shot_assist.
    """
    d = _add_cumulative_mins(df)
    passes = d[d['type.displayName'] == 'Pass'].copy()
    passes['qn'] = passes['qualifiers_parsed'].apply(qual_names)
    passes['completed'] = passes['outcomeType.displayName'] == 'Successful'
    passes['forward'] = passes['endX'] > passes['x']
    passes['headed'] = passes['qn'].apply(lambda s: 'HeadPass' in s)
    passes['is_cross'] = passes['qn'].apply(lambda s: 'Cross' in s)
    passes['is_open_play'] = ~passes['qn'].apply(lambda s: bool(OPEN_PLAY_EXCLUDE_QUALIFIERS & s))
    passes['open_play_cross'] = passes['is_cross'] & passes['is_open_play']
    passes['open_play_cross_completed'] = passes['open_play_cross'] & passes['completed']
    passes['into_final_third'] = (passes['completed'] & passes['is_open_play']
                                   & (passes['x'] < MIDDLE_THIRD_MAX) & (passes['endX'] >= MIDDLE_THIRD_MAX))
    passes['into_box'] = (passes['completed'] & passes['is_open_play']
                           & (~in_box(passes['x'], passes['y'])) & in_box(passes['endX'], passes['endY']))
    passes['shot_assist'] = passes['qn'].apply(lambda s: 'ShotAssist' in s)
    return passes


def compute_passing(df, player_totals, sca_out):
    """
    Per-player passing totals, one table per team. 'Open play' crosses/entries
    exclude corners and free kicks (direct or indirect) - see qualifier sets
    above (see _classify_passes() for the shared classification logic).
    Progressive Passes here is merged in from compute_progressive_passes
    (same definition/logic used elsewhere in this workbook); Shot Assists is
    read directly from WhoScored's own 'ShotAssist' qualifier on the pass
    event, rather than re-derived from the SCA search. SCA is the combined
    count of times a player appears as either the SCA1 or SCA2 contributing
    action in the Shot Creating Actions tab (same underlying search/logic,
    just totalled per player here).
    """
    passes = _classify_passes(df)

    grouped = passes.groupby(['team', 'playerName']).agg(
        passes_completed=('completed', 'sum'),
        passes_attempted=('completed', 'count'),
        passes_forward=('forward', 'sum'),
        headed=('headed', 'sum'),
        crosses_attempted=('open_play_cross', 'sum'),
        crosses_completed=('open_play_cross_completed', 'sum'),
        into_final_third=('into_final_third', 'sum'),
        into_box=('into_box', 'sum'),
        shot_assists=('shot_assist', 'sum'),
    ).reset_index().rename(columns={'playerName': 'player'})

    prog = player_totals[['team', 'player', 'progressive_passes']]
    grouped = grouped.merge(prog, on=['team', 'player'], how='left')
    grouped['progressive_passes'] = grouped['progressive_passes'].fillna(0).astype(int)

    # SCA total = combined count of SCA1 + SCA2 credits from the Shot
    # Creating Actions search. Build a player -> team lookup from the full
    # event log (not just passers) so a player is grouped under their own
    # team even in the rare case a credited action is an opponent's foul.
    player_team_map = (
        df.dropna(subset=['playerName', 'team'])
        .drop_duplicates('playerName')
        .set_index('playerName')['team']
        .to_dict()
    )
    sca_counts = {}
    for col in ['sca1_player', 'sca2_player']:
        if col in sca_out.columns:
            for player, cnt in sca_out[col].dropna().value_counts().items():
                sca_counts[player] = sca_counts.get(player, 0) + int(cnt)
    sca_df = pd.DataFrame([
        {'team': player_team_map.get(p), 'player': p, 'SCA': c}
        for p, c in sca_counts.items()
    ]) if sca_counts else pd.DataFrame(columns=['team', 'player', 'SCA'])

    grouped = grouped.merge(sca_df, on=['team', 'player'], how='outer')
    numeric_cols = ['passes_completed', 'passes_attempted', 'passes_forward', 'headed',
                     'crosses_attempted', 'crosses_completed', 'into_final_third', 'into_box',
                     'shot_assists', 'progressive_passes', 'SCA']
    for c in numeric_cols:
        if c in grouped.columns:
            grouped[c] = grouped[c].fillna(0).astype(int)

    grouped = grouped.rename(columns={
        'passes_completed': 'Passes Completed',
        'passes_attempted': 'Passes Attempted',
        'passes_forward': 'Passes Forward',
        'headed': 'Headed',
        'crosses_attempted': 'Crosses Attempted',
        'crosses_completed': 'Crosses Completed',
        'into_final_third': 'Passes into Final 1/3',
        'into_box': 'Passes into the Box',
        'progressive_passes': 'Progressive Passes',
        'shot_assists': 'Shot Assists',
    })
    grouped = grouped[[
        'team', 'player', 'Passes Completed', 'Passes Attempted', 'Passes Forward', 'Headed',
        'Crosses Attempted', 'Crosses Completed', 'Passes into Final 1/3', 'Passes into the Box',
        'Progressive Passes', 'Shot Assists', 'SCA',
    ]]
    grouped = grouped.sort_values(['team', 'Passes Attempted'], ascending=[True, False])

    return grouped


# ============================================================
# 5c. POSSESSION SEQUENCES (10+ pass chains)
# ============================================================
SEQUENCE_MIN_PASSES = 10
SEQUENCE_ADMIN_EXCLUDE_TYPES = {'OffsideGiven', 'CornerAwarded', 'Card', 'FormationChange',
                                 'Start', 'SubstitutionOff', 'SubstitutionOn', 'FormationSet', 'End'}
# These two parameters (window size, required same-team run within it) were
# tuned empirically against a published benchmark (insight90.streamlit.app's
# "Passes per Sequence" / "10+ Pass Sequences" chart for this same match) -
# see compute_sequences() docstring.
SEQUENCE_CHAIN_CHECK = 5
SEQUENCE_SUC_EVTS_IN_CHAIN = 2


def compute_sequences(df, min_passes=SEQUENCE_MIN_PASSES,
                       chain_check=SEQUENCE_CHAIN_CHECK, suc_evts_in_chain=SEQUENCE_SUC_EVTS_IN_CHAIN):
    """
    Possession sequences per team, using the same windowed possession-chain
    algorithm found in the open-source WhoScored report notebooks behind
    insight90.streamlit.app (a "get_possession_chains" style approach,
    originally from jakeyk11's tutorials). Rather than ending a sequence the
    instant the opponent's team touches the ball, this looks at a sliding
    window of the next `chain_check - 1` events: possession only flips to
    the other team once at least `suc_evts_in_chain` of those events belong
    to them - so a single missed tackle, aerial, or loose touch doesn't end
    a sequence, but a sustained spell by the other side does. Sequences also
    always restart at goals and at the start of a new period (half/extra
    time). `chain_check=5, suc_evts_in_chain=2` were chosen by testing
    several combinations against insight90's published numbers for this
    match (Man Utd 25 sequences of 10+ passes, average 9 passes/sequence;
    Crystal Palace 10 sequences of 10+, average 6) - this combination came
    closest (24/13 sequences, 9.69/5.97 average) without an exact source to
    match bit-for-bit. "Passes" here counts ALL pass attempts in a sequence,
    not just completed ones - that was also the closest-matching definition.
    """
    work = df[~df['type.displayName'].isin(SEQUENCE_ADMIN_EXCLUDE_TYPES)].copy()
    work = work.sort_index().reset_index(drop=True)

    teams_sorted = sorted(work['team'].dropna().unique())
    min_team = teams_sorted[0] if teams_sorted else None
    work['team_binary'] = (work['team'] == min_team).astype(int)
    work['goal_binary'] = (work['type.displayName'] == 'Goal').astype(int).diff().apply(
        lambda x: 1 if x < 0 else 0)
    work['period_num'] = work['period.displayName'].map(PERIOD_ORDER)

    chain = pd.DataFrame(index=work.index)
    same_team_cols = []
    for n in range(1, chain_check):
        col = f'evt_{n}_same_team'
        same_team_cols.append(col)
        chain[col] = abs(work['team_binary'].diff(periods=-n))
        chain[col] = chain[col].apply(lambda x: 1 if x > 1 else x)
    chain['enough_evt_same_team'] = chain[same_team_cols].sum(axis=1).apply(
        lambda x: 1 if x < chain_check - suc_evts_in_chain else 0)
    chain['enough_evt_same_team'] = chain['enough_evt_same_team'].diff(periods=1)
    chain.loc[chain['enough_evt_same_team'] < 0, :] = 0

    chain['upcoming_ko'] = 0
    ko_mask = (work['goal_binary'] == 1) | (work['period_num'].diff().fillna(0) != 0)
    for ko_pos in work.index[ko_mask]:
        lo = max(0, ko_pos - suc_evts_in_chain)
        chain.iloc[lo:ko_pos, chain.columns.get_loc('upcoming_ko')] = 1

    chain['valid_pos_start'] = chain['enough_evt_same_team'].fillna(0) - chain['upcoming_ko'].fillna(0)
    chain['kick_off_period_change'] = work['period_num'].diff(periods=1)
    chain['kick_off_goal'] = work['goal_binary']
    chain.loc[chain['kick_off_period_change'] == 1, 'valid_pos_start'] = 1
    chain.loc[chain['kick_off_goal'] == 1, 'valid_pos_start'] = 1

    chain['team'] = work['team']
    if len(chain) == 0:
        return pd.DataFrame(columns=['team', 'passes']), pd.Series(dtype=int)

    first_idx = chain.index[0]
    chain.loc[first_idx, 'valid_pos_start'] = 1
    chain.loc[first_idx, 'possession_id'] = 1
    chain.loc[first_idx, 'possession_team'] = chain.loc[first_idx, 'team']

    valid_starts = chain.index[chain['valid_pos_start'] > 0]
    possession_id = 2
    for i in range(1, len(valid_starts)):
        cur_pos, prev_pos = valid_starts[i], valid_starts[i - 1]
        current_team = chain.loc[cur_pos, 'team']
        previous_team = chain.loc[prev_pos, 'team']
        if (previous_team == current_team and chain.loc[cur_pos, 'kick_off_goal'] != 1
                and chain.loc[cur_pos, 'kick_off_period_change'] != 1):
            chain.loc[cur_pos, 'possession_id'] = np.nan
        else:
            chain.loc[cur_pos, 'possession_id'] = possession_id
            chain.loc[cur_pos, 'possession_team'] = current_team
            possession_id += 1

    work['possession_id'] = chain['possession_id']
    work['possession_team'] = chain['possession_team']
    work[['possession_id', 'possession_team']] = work[['possession_id', 'possession_team']].ffill().bfill()

    pass_mask = work['type.displayName'] == 'Pass'
    passes_per_chain = work[pass_mask].groupby('possession_id').size()
    chain_team = work.groupby('possession_id')['possession_team'].first()
    chains_df = pd.DataFrame({'passes': passes_per_chain}).reindex(chain_team.index).fillna(0)
    chains_df['team'] = chain_team
    chains_df = chains_df.reset_index(drop=True)

    if len(chains_df):
        long_chains = chains_df[chains_df['passes'] >= min_passes]
        team_sequences = long_chains.groupby('team').size()
    else:
        team_sequences = pd.Series(dtype=int)

    return chains_df, team_sequences


# ============================================================
# 5d. FIELD TILT & PPDA
# ============================================================
PPDA_CUTOFF = 30  # % of pitch; excludes each team's own defensive 30% from PPDA
PPDA_ACTION_TYPES = {'Tackle', 'Interception', 'Foul', 'Challenge'}


def compute_field_tilt(team_summary):
    """
    Field Tilt: each team's share of the two teams' combined touches in
    their attacking (final) third - i.e. touches_in_final_third / (both
    teams' touches_in_final_third), as a percentage. Matches insight90's
    published Field Tilt for this match almost exactly (75.5% vs a
    published 76% for Man Utd, 24.5% vs 24% for Crystal Palace).
    """
    final_third_touches = team_summary.set_index('team')['Final third']
    total = final_third_touches.sum()
    if total == 0:
        return (final_third_touches * 0).rename('Field Tilt %')
    return (final_third_touches / total * 100).round(1).rename('Field Tilt %')


def compute_ppda(df, cutoff=PPDA_CUTOFF, action_types=PPDA_ACTION_TYPES):
    """
    PPDA (Passes Per Defensive Action): how many passes a team allows its
    opponent to complete, on average, before making a defensive action -
    lower means more aggressive pressing. Excludes both the opponent's
    passes AND the pressing team's own defensive actions when they occur
    in that team's own deepest 30% of the pitch (buildup/pressing that deep
    isn't meaningful for pressing intensity). Defensive actions counted:
    Tackle, Interception, Foul, Challenge (any outcome) - Aerials and
    Clearances are excluded, the standard convention for this stat. Tuned
    against insight90's published PPDA for this match (6.92 for Man Utd,
    15.42 for Crystal Palace) - this gets close (6.6 / 15.3) but not exact,
    since the source implementation isn't public.
    """
    teams = [t for t in df['team'].dropna().unique()]
    result = {}
    for t in teams:
        others = [x for x in teams if x != t]
        if not others:
            continue
        opp = others[0]
        opp_passes = df[(df['team'] == opp) & (df['type.displayName'] == 'Pass') & (df['x'] >= cutoff)].shape[0]
        own_def_actions = df[(df['team'] == t) & (df['type.displayName'].isin(action_types))
                              & (df['x'] >= cutoff)].shape[0]
        result[t] = round(opp_passes / own_def_actions, 2) if own_def_actions else None
    return pd.Series(result, name='PPDA')


# ============================================================
# 5d2. DEFENSIVE STATS
# ============================================================
# Formula history for Defensive Action Height, since it's an approximation
# (insight90's real source isn't public) that's now been tuned against TWO
# real matches instead of one, via diagnose_def_action_height.py:
#   v1 (original): median x of {Tackle, Clearance, BallRecovery, Challenge,
#       Aerial} - fit ONE match well (41.42m vs published 41.44m for Man
#       Utd, 31.4m vs 31.81m for Crystal Palace) but fell apart on a second
#       match (Man Utd vs Brentford: 21.21m vs a published 26.55m, 36.44m
#       vs 38.47m) - a classic overfit-to-one-example failure.
#   v2 (current): MEAN (not median) x of {Tackle, Interception, Clearance,
#       BallRecovery, Challenge} - Aerial dropped entirely (it was actively
#       hurting the fit: including it pushed the combined error on the
#       Crystal Palace match up to 10+ metres in testing), goalkeeper events
#       excluded (see _guess_goalkeepers() below). This combination won a
#       grid search across 22 candidate formulas run against BOTH matches:
#       combined error 3.88m total (Brentford: 25.96/36.55 vs 26.55/38.47;
#       Crystal Palace: 41.61/33.01 vs 41.44/31.81) - noticeably tighter
#       than v1's fit once you require it to work on more than one match.
DEF_ACTION_HEIGHT_TYPES = {'Tackle', 'Interception', 'Clearance', 'BallRecovery', 'Challenge'}


def _guess_goalkeepers(df):
    """
    Best-effort per-team goalkeeper identification, used to exclude GK
    events from Defensive Action Height below (a team's average defensive
    LINE height is meant to represent the outfield block, not wherever the
    keeper happens to touch the ball - and a keeper sweeping/clearing very
    close to their own goal would otherwise drag a team's number down for
    reasons that have nothing to do with their outfield defenders' shape).

    WhoScored's own event stream carries no explicit position field, so
    this infers it indirectly: whichever player recorded the most 'Save'
    events on a team is almost certainly that team's keeper (outfield
    players essentially never record a Save). If a keeper somehow made zero
    saves in a match, this returns no answer for that team and every event
    is included for them instead of guessing wrong - a graceful fallback,
    not a crash. Verified against two real matches (correctly identified
    both keepers in each) as part of tuning this formula - see
    diagnose_def_action_height.py's own version of this same logic.
    """
    saves = df[df['type.displayName'] == 'Save']
    gks = {}
    for team, grp in saves.groupby('team'):
        counts = grp['playerName'].value_counts()
        if not counts.empty:
            gks[team] = counts.index[0]
    return gks


def compute_defensive_stats(df):
    """
    Six defensive counting stats per team, all read directly from discrete
    WhoScored event types/qualifiers (no heuristics needed, unlike PPDA/
    sequences) - verified to match insight90's published numbers for this
    match exactly: Tackles/Successful Tackles/Interceptions come straight
    from Tackle and Interception event counts; Blocked Passes counts
    'BlockedPass' events (the blocking team's own event); Blocked Shots
    counts the OPPONENT's shot attempts carrying a 'Blocked' qualifier,
    credited to this team (the side that did the blocking). Defensive
    Action Height is the MEAN x position (converted to metres) of this
    team's Tackle/Interception/Clearance/BallRecovery/Challenge events,
    EXCLUDING that team's own goalkeeper (see _guess_goalkeepers()) - the
    one metric here that's tuned/approximate rather than an exact event
    count, since the source's precise definition isn't public. This
    specific formula (mean, no Aerial, no GK) was chosen by grid-searching
    22 candidate formulas against two real matches at once rather than one -
    see the version-history comment above DEF_ACTION_HEIGHT_TYPES for the
    exact numbers, and diagnose_def_action_height.py for the search itself.
    """
    teams = [t for t in df['team'].dropna().unique()]

    tackles = df[df['type.displayName'] == 'Tackle'].groupby('team').size()
    succ_tackles = df[(df['type.displayName'] == 'Tackle')
                       & (df['outcomeType.displayName'] == 'Successful')].groupby('team').size()
    interceptions = df[df['type.displayName'] == 'Interception'].groupby('team').size()
    blocked_passes = df[df['type.displayName'] == 'BlockedPass'].groupby('team').size()

    blocked_shots = {}
    shots = df[df['isShot'] == True].copy()
    shots['qn'] = shots['qualifiers_parsed'].apply(qual_names)
    shots_blocked = shots[shots['qn'].apply(lambda s: 'Blocked' in s)]
    for t in teams:
        others = [x for x in teams if x != t]
        opp = others[0] if others else None
        blocked_shots[t] = shots_blocked[shots_blocked['team'] == opp].shape[0] if opp else 0

    da = df[df['type.displayName'].isin(DEF_ACTION_HEIGHT_TYPES)]
    gk_names = _guess_goalkeepers(df)
    da = da[~da.apply(lambda r: gk_names.get(r['team']) == r['playerName'], axis=1)]
    def_action_height = (da.groupby('team')['x'].mean() * (PITCH_LEN_M / 100)).round(2)

    out = pd.DataFrame({
        'Tackles': tackles,
        'Successful Tackles': succ_tackles,
        'Interceptions': interceptions,
        'Blocked Passes': blocked_passes,
    })
    out['Blocked Shots'] = pd.Series(blocked_shots)
    out['Defensive Action Height (m)'] = def_action_height
    return out


def compute_corners(df):
    """Total corner kicks taken per team - passes carrying the 'CornerTaken' qualifier."""
    qn = df['qualifiers_parsed'].apply(qual_names)
    corners = df[qn.apply(lambda s: 'CornerTaken' in s)]
    return corners.groupby('team').size().rename('Corners')


def compute_defensive_actions(df):
    """
    Per-player totals for the Defensive Actions tab, one table per team:
    Tackles and Interceptions are direct event counts; Passes Blocked
    counts 'BlockedPass' events by their own playerName (the blocker, not
    the passer). Shots Blocked has no direct player field in the data - a
    blocked shot only carries a 'Blocked' qualifier on the SHOOTER's own
    event, with no separate row for the blocking defender - but WhoScored
    pairs every blocked shot with a 'Save' event on the very next row,
    credited to the defender (or keeper) who actually stopped it; that
    Save event's player is used as the blocker here. This was verified to
    sum to the exact same team totals as compute_defensive_stats() (2 for
    Man Utd, 5 for Crystal Palace in this match). Every player who appears
    anywhere in the match gets a row (zeros where they had none of these
    four actions), matching the Passing/Touches tabs' inclusiveness.
    """
    all_players = (df.dropna(subset=['playerName', 'team'])
                    .drop_duplicates(['team', 'playerName'])[['team', 'playerName']]
                    .rename(columns={'playerName': 'player'}))

    tackles = (df[df['type.displayName'] == 'Tackle'].groupby(['team', 'playerName']).size()
               .reset_index(name='Tackles').rename(columns={'playerName': 'player'}))
    interceptions = (df[df['type.displayName'] == 'Interception'].groupby(['team', 'playerName']).size()
                      .reset_index(name='Interceptions').rename(columns={'playerName': 'player'}))
    passes_blocked = (df[df['type.displayName'] == 'BlockedPass'].groupby(['team', 'playerName']).size()
                       .reset_index(name='Passes Blocked').rename(columns={'playerName': 'player'}))

    work = df.sort_index().reset_index(drop=True)
    shots = work[work['isShot'] == True].copy()
    shots['qn'] = shots['qualifiers_parsed'].apply(qual_names)
    blocked_idx = shots[shots['qn'].apply(lambda s: 'Blocked' in s)].index
    blocker_rows = []
    for idx in blocked_idx:
        if idx + 1 < len(work):
            nxt = work.iloc[idx + 1]
            if nxt['type.displayName'] == 'Save' and pd.notna(nxt.get('team')) and pd.notna(nxt.get('playerName')):
                blocker_rows.append({'team': nxt['team'], 'player': nxt['playerName']})
    shots_blocked = (pd.DataFrame(blocker_rows).groupby(['team', 'player']).size()
                      .reset_index(name='Shots Blocked')) if blocker_rows else pd.DataFrame(
        columns=['team', 'player', 'Shots Blocked'])

    out = all_players
    for piece in [tackles, interceptions, passes_blocked, shots_blocked]:
        out = out.merge(piece, on=['team', 'player'], how='left')
    for c in ['Tackles', 'Interceptions', 'Passes Blocked', 'Shots Blocked']:
        out[c] = out[c].fillna(0).astype(int)
    out = out.sort_values(['team', 'Tackles'], ascending=[True, False])
    return out


# ============================================================
# 5d3. DEFENSIVE ACTION LOCATION (per player, by pitch third)
# ============================================================
DEF_ACTION_LOCATION_METRICS = [
    ('Tackle', 'Tackles'),
    ('Interception', 'Interceptions'),
    ('BlockedPass', 'Passes Blocked'),
    ('BallRecovery', 'Ball Recoveries'),
]
DEF_ACTION_LOCATION_THIRDS = [
    ('Own third', 'Own Third'),
    ('Middle third', 'Middle Third'),
    ('Final third', 'Final Third'),
]


def compute_defensive_action_location(df):
    """
    Per-player totals for the Defensive Action Location tab, one table per
    team: Tackles, Interceptions, Passes Blocked, and Ball Recoveries, each
    broken down by which third of the pitch the action happened in - using
    the exact same Own/Middle/Final third rule as the Touches tab (third(),
    based on the event's own x coordinate, 0-100 normalized scale, each
    team's own attacking direction). Passes Blocked uses the blocking
    player's own event (same convention as the Defensive Actions tab), not
    the passer's. Every player who appears anywhere in the match gets a row
    (zeros where they had none), matching the other tabs' inclusiveness.
    """
    all_players = (df.dropna(subset=['playerName', 'team'])
                    .drop_duplicates(['team', 'playerName'])[['team', 'playerName']]
                    .rename(columns={'playerName': 'player'}))

    out = all_players
    value_cols = []
    for event_type, label in DEF_ACTION_LOCATION_METRICS:
        work = df[df['type.displayName'] == event_type].copy()
        work['pitch_third'] = work['x'].apply(third)
        for third_label, col_suffix in DEF_ACTION_LOCATION_THIRDS:
            col_name = f'{label} {col_suffix}'
            value_cols.append(col_name)
            counts = (work[work['pitch_third'] == third_label].groupby(['team', 'playerName']).size()
                      .reset_index(name=col_name).rename(columns={'playerName': 'player'}))
            out = out.merge(counts, on=['team', 'player'], how='left')

    for c in value_cols:
        out[c] = out[c].fillna(0).astype(int)

    out['_sort_key'] = out[value_cols].sum(axis=1)
    out = out.sort_values(['team', '_sort_key'], ascending=[True, False]).drop(columns='_sort_key')
    out = out[['team', 'player'] + value_cols].reset_index(drop=True)
    return out


# ============================================================
# 5e. TOTALS (team-level rollup)
# ============================================================
def compute_totals(team_summary, team_totals, passing_out, sca_out, chains_df, team_sequences,
                    field_tilt, ppda, defensive_stats, corners, home_name, away_name):
    """
    One row per team, rolling up numbers already computed elsewhere:
    Shots = count of shot events (rows in sca_out); Possession % is each
    team's share of total passes attempted (team passes / both teams'
    passes x 100) - a standard passes-based possession proxy; Touches/thirds/
    box and carry numbers come from the Touches tab's team summary; Total
    Passes, Crosses, Passes into Final Third/Box, and Total SCA are summed
    from the Passing tab's per-player numbers; Progressive Passes matches
    the Passing tab's team-level total for that column; 10+ Pass Sequences and Avg
    Passes per Sequence both come from compute_sequences() - the latter is
    the mean passes-per-chain across ALL possession sequences (not just the
    10+ ones), using the exact same sequence definition; Field Tilt and PPDA
    come from compute_field_tilt()/compute_ppda(); Tackles/Successful
    Tackles/Interceptions/Blocked Passes/Blocked Shots/Defensive Action
    Height come from compute_defensive_stats(); Corners is the count of Pass
    events carrying the 'CornerTaken' qualifier, from compute_corners().
    """
    shots = sca_out.groupby('team').size().rename('Shots')

    base = team_summary.set_index('team')[[
        'Total touches', 'Own third', 'Middle third', 'Final third', 'Attacking Box',
        'Progressive Carries', 'Carries into Final Third', 'Carries into Box',
    ]].rename(columns={
        'Total touches': 'Touches',
        'Own third': 'Touches in Own Third',
        'Middle third': 'Touches in Middle Third',
        'Final third': 'Touches in Attacking Third',
        'Attacking Box': "Touches in Opponent's Box",
        'Carries into Box': "Carries into Opponent's Box",
    })

    passing_team = passing_out.groupby('team').agg(**{
        'Total Passes': ('Passes Attempted', 'sum'),
        'Crosses Attempted': ('Crosses Attempted', 'sum'),
        'Crosses Completed': ('Crosses Completed', 'sum'),
        'Passes into Final Third': ('Passes into Final 1/3', 'sum'),
        "Passes into Opponent's Box": ('Passes into the Box', 'sum'),
        'Total SCA': ('SCA', 'sum'),
    })

    prog = team_totals.set_index('team')[['progressive_passes']].rename(
        columns={'progressive_passes': 'Progressive Passes'})

    def_counts = defensive_stats[['Tackles', 'Successful Tackles', 'Interceptions',
                                   'Blocked Passes', 'Blocked Shots']]

    # Possession % is passes-based: each team's share of the combined pass count.
    total_passes_for_poss = passing_team['Total Passes'].fillna(0)
    possession = (total_passes_for_poss / total_passes_for_poss.sum() * 100).round(1)

    totals = base.join(passing_team, how='outer').join(prog, how='outer').join(def_counts, how='outer')
    totals['Shots'] = shots
    totals['10+ Pass Sequences'] = team_sequences
    totals['Corners'] = corners
    totals = totals.fillna(0)
    for c in totals.columns:
        totals[c] = totals[c].astype(int)

    avg_passes = chains_df.groupby('team')['passes'].mean().round(2) if len(chains_df) else pd.Series(dtype=float)
    totals['Avg Passes per Sequence'] = avg_passes
    totals['Avg Passes per Sequence'] = totals['Avg Passes per Sequence'].fillna(0.0)
    totals['Field Tilt %'] = field_tilt
    totals['Field Tilt %'] = totals['Field Tilt %'].fillna(0.0)
    totals['PPDA'] = ppda
    totals['PPDA'] = totals['PPDA'].fillna(0.0)
    totals['Defensive Action Height (m)'] = defensive_stats['Defensive Action Height (m)']
    totals['Defensive Action Height (m)'] = totals['Defensive Action Height (m)'].fillna(0.0)
    totals['Possession %'] = possession
    totals['Possession %'] = totals['Possession %'].fillna(0.0)

    teams_order = [t for t in [home_name, away_name] if t is not None] or sorted(totals.index)
    totals = totals.reindex(teams_order)
    totals = totals.reset_index()

    totals = totals[[
        'team', 'Shots', 'Possession %', 'Total SCA', 'Touches', 'Touches in Own Third',
        'Touches in Middle Third', 'Touches in Attacking Third', "Touches in Opponent's Box",
        'Progressive Carries', 'Carries into Final Third', "Carries into Opponent's Box",
        'Total Passes', 'Crosses Attempted', 'Crosses Completed',
        'Passes into Final Third', "Passes into Opponent's Box",
        'Progressive Passes', '10+ Pass Sequences', 'Avg Passes per Sequence',
        'Tackles', 'Successful Tackles', 'Interceptions', 'Blocked Passes', 'Blocked Shots',
        'Field Tilt %', 'PPDA',
        'Defensive Action Height (m)', 'Corners',
    ]]
    return totals


def compute_against_totals(totals_out):
    """
    'Against' tab: the same columns as the Totals tab, but every stat is the
    OPPONENT's number rather than the team's own - e.g. the 'vs Man Utd' row's
    Shots value is Crystal Palace's total shots, since that's how many shots
    were taken AGAINST Man Utd. Only meaningful (and only supported) for the
    standard 2-team match case.
    """
    if len(totals_out) != 2:
        raise ValueError("compute_against_totals expects exactly 2 teams in totals_out")
    against = totals_out.iloc[::-1].reset_index(drop=True).copy()
    against['team'] = totals_out['team'].apply(lambda t: f'vs {t}').values
    return against


# ============================================================
# 5f. ON/OFF (per-player on-pitch splits)
# ============================================================
def extract_player_windows(df):
    """
    Per-player on-pitch window for this match: Team, Player, start_minute,
    end_minute, Minutes Played, subbed_off. Backs the On/Off tab (see
    compute_on_off() below) - the WhoScored-side equivalent of
    fotmob_report.extract_player_windows(), same underlying concept (a
    player's own on-pitch window) but read directly from WhoScored's own
    event stream instead of FotMob's lineup JSON. WhoScored's own
    'SubstitutionOn'/'SubstitutionOff' events already carry the exact
    minute (via cumulative match time, continuous across halves - see
    _add_cumulative_mins()) each substitution happened, so unlike the
    FotMob version, there's no need for a "Minutes Played" fallback
    estimate - the real event log always has it.

    start_minute is a player's own 'SubstitutionOn' time if they came on
    as a substitute, else 0 (they started the match). end_minute is their
    own 'SubstitutionOff' time if they were substituted off, else the
    match's own last recorded event time (correctly includes stoppage
    time in both halves, since it's read from the real event log rather
    than assumed to be a flat 90/45).

    subbed_off (bool) tells compute_on_off() how to treat the boundary at
    end_minute, the same convention fotmob_report.compute_plus_minus()
    uses: a substituted player doesn't get credit for the exact minute
    they left (that belongs to whoever replaced them), while a player who
    finished the match does get the exact final minute.

    KNOWN LIMITATION: a sent-off player (red card / second yellow) isn't
    detected here - their window still runs to the final whistle rather
    than ending at the card. See compute_on_off()'s own docstring for what
    that means for its "Against" numbers on a match with a red card in it.

    Only players with at least one event of their own in this match's
    event stream get a row - a bench player who never actually appeared
    naturally has none. (In the very rare case of an outfield player who
    touched the ball zero times all match - no pass, no tackle, nothing -
    they'd be missing here too; every other per-player table in this
    workbook has this same limitation, since they all key off
    df['playerName'] the same way.)
    """
    columns = ['Team', 'Player', 'start_minute', 'end_minute', 'Minutes Played', 'subbed_off']
    if df.empty:
        return pd.DataFrame(columns=columns)

    d = _add_cumulative_mins(df)
    match_end = d['cumulative_mins'].max()

    players = d[d['playerName'].notna()][['team', 'playerName']].drop_duplicates()

    sub_off = d[d['type.displayName'] == 'SubstitutionOff']
    off_lookup = {(r.team, r.playerName): r.cumulative_mins for r in sub_off.itertuples()}

    sub_on = d[d['type.displayName'] == 'SubstitutionOn']
    on_lookup = {(r.team, r.playerName): r.cumulative_mins for r in sub_on.itertuples()}

    rows = []
    for r in players.itertuples():
        team, player = r.team, r.playerName
        start = on_lookup.get((team, player), 0.0)
        subbed_off = (team, player) in off_lookup
        end = off_lookup[(team, player)] if subbed_off else match_end
        rows.append({
            'Team': team, 'Player': player,
            'start_minute': start, 'end_minute': end,
            'Minutes Played': round(end - start),
            'subbed_off': subbed_off,
        })

    return pd.DataFrame(rows, columns=columns)


def compute_on_off(df, player_windows, carries_df):
    """
    On/Off tab: for every player, team totals - Shots, Total Touches, Own/
    Middle/Final Third Touches, Attacking Box Touches, Carries into Final
    Third, Carries into Box, Passes into Final 1/3, Passes into the Box,
    Crosses Attempted, and Crosses Completed - split into "For" (their own
    team's total) and "Against" (the opponent's total), counted only for
    the minutes that player was actually on the pitch (their own window
    from extract_player_windows() above) - the same on-pitch-window
    concept fotmob_report.py's Plus Minus tab uses for FotMob's shot/goal/
    xG data, applied here to WhoScored's own event stats instead. All 12
    metrics' "For" columns are grouped together, followed by all 12
    "Against" columns, rather than interleaved pair by pair.

    Three more metrics - Interceptions, Blocked Passes, and Blocked Shots -
    are also included, but "For" only (no "Against" pair): these are
    already inherently one-sided defensive actions (my team intercepting/
    blocking something the opponent did), so an "Against" version would be
    a materially different stat (how often the OPPONENT blocked/
    intercepted THIS team), which wasn't asked for here. Interceptions and
    Blocked Passes are direct event-type counts (WhoScored's own
    'Interception'/'BlockedPass' events, credited to the team that made
    the play); Blocked Shots uses the exact same definition as the Totals
    tab's own Blocked Shots column (compute_defensive_stats()) - the
    OPPONENT's shot attempts carrying a 'Blocked' qualifier, credited to
    this team as the side that did the blocking - rather than the more
    involved Save-event-pairing compute_defensive_actions() needs to name
    a specific blocking PLAYER, since a team total doesn't need to know
    who exactly made the block.

    Own Third / Middle Third / Final Third / Attacking Box / Carries into
    Final Third / Carries into Box use the exact same thirds boundaries and
    box definition as the Touches tab (OWN_THIRD_MAX/MIDDLE_THIRD_MAX,
    in_box()); Passes into Final 1/3 / Passes into the Box / Crosses
    Attempted / Crosses Completed use the exact same open-play-only
    definitions as the Passing tab (see _classify_passes()) - one shared
    implementation for each, so this tab can never quietly disagree with
    the Touches/Passing tabs about what counts. Like every other thirds/
    box stat in this project, that's each event's OWN attacking direction
    (WhoScored/Opta coordinates already run 0 at a team's own goal to 100
    at the opponent's goal for every event, regardless of which actual
    half of the pitch or which way that team is actually shooting) - so
    e.g. "Final Third Against" means the OPPONENT was in THEIR OWN
    attacking final third (i.e. deep in this player's own defensive
    third), not literally the same physical strip of grass as "Final
    Third For".

    carries_df is compute_carries()'s own raw, one-row-per-carry data (see
    that function's docstring) - needed here because carries aren't a
    single WhoScored event row with their own minute the way touches/
    shots/passes are; each carry's cumulative_mins (when it started) is
    what's used to decide which player's on-pitch window it falls in.

    A player who was subbed off doesn't get credit for the exact minute
    they left (that belongs to whoever replaced them) - a player who
    played to the final whistle does get credit for the exact final
    minute. Same boundary convention as fotmob_report.compute_plus_minus().

    KNOWN LIMITATION: a red/second-yellow card isn't accounted for - a
    sent-off player's own window still runs to the final whistle rather
    than ending at the card, and this project has no explicit per-minute
    "how many players were on the pitch" tracking, only each individual
    player's own window - so a match with a red card in it will somewhat
    understate how much the shorthanded team was actually under the cosh
    for those minutes. Not expected to come up often, but worth knowing
    before drawing conclusions from a match that had a sending-off in it.

    Returns one DataFrame, sorted by Team then descending Minutes Played -
    split into two tables, one per team, when written to the
    workbook/UI (see build_workbook()).
    """
    for_cols = ['Shots For', 'Total Touches For', 'Own Third For', 'Middle Third For',
                'Final Third For', 'Attacking Box For', 'Carries into Final Third For',
                'Carries into Box For', 'Passes into Final 1/3 For', 'Passes into the Box For',
                'Crosses Attempted For', 'Crosses Completed For']
    against_cols = [c.replace(' For', ' Against') for c in for_cols]
    for_only_cols = ['Interceptions For', 'Blocked Passes For', 'Blocked Shots For']
    out_cols = ['Team', 'Player', 'Minutes Played'] + for_cols + for_only_cols + against_cols
    if df.empty or player_windows.empty:
        return pd.DataFrame(columns=out_cols)

    d = _add_cumulative_mins(df)
    d['pitch_third'] = d['x'].apply(third)
    d['in_att_box'] = in_box(d['x'], d['y'])

    passes = _classify_passes(df)

    # Blocked Shots For: the OPPONENT's own shot attempts carrying a
    # 'Blocked' qualifier - same definition as compute_defensive_stats()'s
    # Totals-tab Blocked Shots column (see this function's own docstring).
    blocked_shot_events = d[d['isShot'] == True].copy()
    blocked_shot_events['qn'] = blocked_shot_events['qualifiers_parsed'].apply(qual_names)
    blocked_shot_events = blocked_shot_events[blocked_shot_events['qn'].apply(lambda s: 'Blocked' in s)]

    teams_all = sorted(player_windows['Team'].dropna().unique())
    opponent = {teams_all[0]: teams_all[1], teams_all[1]: teams_all[0]} if len(teams_all) == 2 else {}

    records = []
    for _, prow in player_windows.iterrows():
        team = prow['Team']
        opp = opponent.get(team)
        start, end, subbed_off = prow['start_minute'], prow['end_minute'], prow['subbed_off']

        if subbed_off:
            in_window = lambda mins: (mins >= start) & (mins < end)
        else:
            in_window = lambda mins: (mins >= start) & (mins <= end)

        mask = in_window(d['cumulative_mins'])
        own = d[(d['team'] == team) & mask]
        opp_events = d[(d['team'] == opp) & mask] if opp else d.iloc[0:0]
        own_touches = own[own['isTouch'] == True]
        opp_touches = opp_events[opp_events['isTouch'] == True]

        interceptions_for = int((own['type.displayName'] == 'Interception').sum())
        blocked_passes_for = int((own['type.displayName'] == 'BlockedPass').sum())
        if opp and not blocked_shot_events.empty:
            bs_mask = in_window(blocked_shot_events['cumulative_mins'])
            blocked_shots_for = int(((blocked_shot_events['team'] == opp) & bs_mask).sum())
        else:
            blocked_shots_for = 0

        pass_mask = in_window(passes['cumulative_mins'])
        own_passes = passes[(passes['team'] == team) & pass_mask]
        opp_passes = passes[(passes['team'] == opp) & pass_mask] if opp else passes.iloc[0:0]

        if carries_df.empty:
            own_carries = opp_carries = carries_df
        else:
            carry_mask = in_window(carries_df['cumulative_mins'])
            own_carries = carries_df[(carries_df['team'] == team) & carry_mask]
            opp_carries = carries_df[(carries_df['team'] == opp) & carry_mask] if opp else carries_df.iloc[0:0]

        records.append({
            'Team': team,
            'Player': prow['Player'],
            'Minutes Played': prow['Minutes Played'],
            'Shots For': int((own['isShot'] == True).sum()),
            'Total Touches For': int(len(own_touches)),
            'Own Third For': int((own_touches['pitch_third'] == 'Own third').sum()),
            'Middle Third For': int((own_touches['pitch_third'] == 'Middle third').sum()),
            'Final Third For': int((own_touches['pitch_third'] == 'Final third').sum()),
            'Attacking Box For': int(own_touches['in_att_box'].sum()),
            'Carries into Final Third For': int(own_carries['into_final_third'].sum()) if len(own_carries) else 0,
            'Carries into Box For': int(own_carries['into_box'].sum()) if len(own_carries) else 0,
            'Passes into Final 1/3 For': int(own_passes['into_final_third'].sum()),
            'Passes into the Box For': int(own_passes['into_box'].sum()),
            'Crosses Attempted For': int(own_passes['open_play_cross'].sum()),
            'Crosses Completed For': int(own_passes['open_play_cross_completed'].sum()),
            'Interceptions For': interceptions_for,
            'Blocked Passes For': blocked_passes_for,
            'Blocked Shots For': blocked_shots_for,
            'Shots Against': int((opp_events['isShot'] == True).sum()),
            'Total Touches Against': int(len(opp_touches)),
            'Own Third Against': int((opp_touches['pitch_third'] == 'Own third').sum()),
            'Middle Third Against': int((opp_touches['pitch_third'] == 'Middle third').sum()),
            'Final Third Against': int((opp_touches['pitch_third'] == 'Final third').sum()),
            'Attacking Box Against': int(opp_touches['in_att_box'].sum()),
            'Carries into Final Third Against': int(opp_carries['into_final_third'].sum()) if len(opp_carries) else 0,
            'Carries into Box Against': int(opp_carries['into_box'].sum()) if len(opp_carries) else 0,
            'Passes into Final 1/3 Against': int(opp_passes['into_final_third'].sum()),
            'Passes into the Box Against': int(opp_passes['into_box'].sum()),
            'Crosses Attempted Against': int(opp_passes['open_play_cross'].sum()),
            'Crosses Completed Against': int(opp_passes['open_play_cross_completed'].sum()),
        })

    out = pd.DataFrame(records, columns=out_cols)
    out = out.sort_values(['Team', 'Minutes Played'], ascending=[True, False]).reset_index(drop=True)
    return out


# ============================================================
# 6. WRITE WORKBOOK
# ============================================================
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
SUBHEADER_FONT = Font(name='Arial', bold=True, color='1F4E78', size=12)
BODY_FONT = Font(name='Arial')


def write_table(ws, df, start_row, start_col=1, title=None):
    r = start_row
    if title:
        c = ws.cell(row=r, column=start_col, value=title)
        c.font = SUBHEADER_FONT
        r += 1
    header_row = r
    for j, col in enumerate(df.columns):
        cell = ws.cell(row=header_row, column=start_col + j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    for i, row in enumerate(df.itertuples(index=False), start=1):
        for j, val in enumerate(row):
            if pd.isna(val):
                val = None
            cell = ws.cell(row=header_row + i, column=start_col + j, value=val)
            cell.font = BODY_FONT
    return header_row + len(df)


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 10), 42)


def build_workbook(sca_out, team_summary, player_third, passing_out, totals_out, defensive_actions,
                    defensive_action_location, passing_pairs, home_name, away_name, against_totals=None,
                    shot_pairs=None, on_off=None):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Totals')
    write_table(ws, totals_out, start_row=1)
    autosize(ws)
    ws.freeze_panes = 'A2'

    if against_totals is not None:
        ws = wb.create_sheet('Against')
        write_table(ws, against_totals, start_row=1)
        autosize(ws)
        ws.freeze_panes = 'A2'

    ws = wb.create_sheet('Touches')
    teams = [t for t in [home_name, away_name] if t is not None] or sorted(player_third['team'].unique())
    r2 = 1
    for t in teams:
        t_players = player_third[player_third['team'] == t].drop(columns=['team'])
        r2 = write_table(ws, t_players, start_row=r2, title=f'{t} - Player Touch Locations') + 3
    autosize(ws)
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Passing')
    teams_for_passing = [t for t in [home_name, away_name] if t is not None] or sorted(passing_out['team'].unique())
    rP = 1
    for t in teams_for_passing:
        t_players = passing_out[passing_out['team'] == t].drop(columns=['team'])
        rP = write_table(ws, t_players, start_row=rP, title=f'{t} - Passing') + 3
    autosize(ws)
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Shot Creating Actions')
    teams_for_sca = [t for t in [home_name, away_name] if t is not None] or sorted(sca_out['team'].unique())
    rSCA = 1
    for t in teams_for_sca:
        t_sca = sca_out[sca_out['team'] == t].drop(columns=['team']).reset_index(drop=True)
        rSCA = write_table(ws, t_sca, start_row=rSCA, title=f'{t} - Shot Creating Actions') + 3
    autosize(ws)
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Defensive Actions')
    teams_for_def = [t for t in [home_name, away_name] if t is not None] or sorted(defensive_actions['team'].unique())
    rD = 1
    for t in teams_for_def:
        t_players = defensive_actions[defensive_actions['team'] == t].drop(columns=['team'])
        rD = write_table(ws, t_players, start_row=rD, title=f'{t} - Defensive Actions') + 3
    autosize(ws)
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Defensive Action Location')
    teams_for_defloc = ([t for t in [home_name, away_name] if t is not None]
                         or sorted(defensive_action_location['team'].unique()))
    rDL = 1
    for t in teams_for_defloc:
        t_players = defensive_action_location[defensive_action_location['team'] == t].drop(columns=['team'])
        rDL = write_table(ws, t_players, start_row=rDL, title=f'{t} - Defensive Action Location') + 3
    autosize(ws)
    ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Passing Pairs')
    teams_for_pairs = [t for t in [home_name, away_name] if t is not None] or sorted(passing_pairs['team'].unique())
    rPP = 1
    for t in teams_for_pairs:
        t_pairs = passing_pairs[passing_pairs['team'] == t].drop(columns=['team'])
        rPP = write_table(ws, t_pairs, start_row=rPP, title=f'{t} - Passing Pairs') + 3
    autosize(ws)
    ws.freeze_panes = 'A3'

    if shot_pairs is not None:
        ws = wb.create_sheet('Shot Pairs')
        teams_for_shot_pairs = ([t for t in [home_name, away_name] if t is not None]
                                 or sorted(shot_pairs['team'].unique()))
        rSP = 1
        for t in teams_for_shot_pairs:
            t_shot_pairs = shot_pairs[shot_pairs['team'] == t].drop(columns=['team'])
            rSP = write_table(ws, t_shot_pairs, start_row=rSP, title=f'{t} - Shot Pairs') + 3
        autosize(ws)
        ws.freeze_panes = 'A3'

    if on_off is not None:
        ws = wb.create_sheet('On-Off')
        teams_for_on_off = ([t for t in [home_name, away_name] if t is not None]
                             or sorted(on_off['Team'].unique()))
        rOO = 1
        for t in teams_for_on_off:
            t_on_off = on_off[on_off['Team'] == t].drop(columns=['Team'])
            rOO = write_table(ws, t_on_off, start_row=rOO, title=f'{t} - On/Off') + 3
        autosize(ws)
        ws.freeze_panes = 'A3'

    ws = wb.create_sheet('Notes')
    notes = [
        "Definitions and assumptions used in this workbook",
        "",
        f"Pitch dimensions assumed: {PITCH_LEN_M}m x {PITCH_WID_M}m (standard Opta convention),"
        " used to convert 0-100 normalized coordinates into yards.",
        "",
        f"Progressive pass: a completed, OPEN PLAY pass that moves the ball at least {PROGRESSIVE_YARD_THRESHOLD} yards"
        f" closer to the opponent's goal than the most-advanced point reached in that team's previous"
        f" {PROGRESSIVE_ROLLING_WINDOW} completed passes (rolling window, not reset by turnovers), OR any"
        " completed, open play pass into the penalty area. Corners, free kicks (direct/indirect), throw-ins,"
        " goal kicks, and keeper throws can never be counted as a progressive pass, though they still feed"
        " into the rolling baseline used to measure another pass's progress. Passes starting in the"
        f" defending {DEFENSIVE_START_CUTOFF}% of the pitch are excluded unless they end in the box.",
        "",
        "Progressive carry: WhoScored/Opta data has no explicit carry event, so carries are inferred using"
        " the same approach as well-known open-source WhoScored report notebooks: for each event, look"
        " ahead to the next one, skipping past TakeOns and Fouls (which don't represent a genuine loss of"
        " the ball), then treat the gap between the first event's end location and that next event's"
        f" start location as a carry - provided both belong to the same team, the first event isn't a bare"
        f" 'BallTouch', the distance covered is between {CARRY_MIN_LEN:.0f} and {CARRY_MAX_LEN:.0f} metres,"
        f" and the elapsed time is between {CARRY_MIN_DUR:.0f} and {CARRY_MAX_DUR:.0f} seconds. A carry"
        f" counts as progressive if it moves the ball at least {PROGRESSIVE_YARD_THRESHOLD} (pitch-length)"
        " units closer to the opponent's goal, or it enters the penalty area. Carries starting in a team's"
        " own third still count, but carries ENDING in their own third are excluded.",
        "",
        "Carries into the final third / into the box: a carry that starts outside that zone and ends"
        " inside it (an entry, not merely occurring within it).",
        "",
        "Shot-creating actions: the up-to-two offensive actions (passes, take-ons, defensive actions"
        " winning the ball, fouls drawn, or shots leading to a rebound shot) immediately before a shot,"
        " walking backward through the same team's play. Aerial duels and ground challenges never break"
        " this search, since WhoScored logs them as separate rows for both sides rather than as genuine"
        " turnovers - but a won aerial does NOT count as a contributing action (it's skipped over so the"
        " search can keep looking further back); a won challenge still counts. An opponent's Clearance,"
        " blocked pass, or keeper Save immediately before the shot is treated the same way - skipped over,"
        " not counted itself - because a defender's partial touch (e.g. heading a cross away, only for it to"
        " fall to another attacker) doesn't erase the real originating pass. Without this, headers set up by"
        " a cross/corner that a defender partially touched first were showing no SCA at all, which was wrong"
        " - someone still played the pass that created the chance. A genuine opponent Tackle, Interception,"
        " or BallRecovery still breaks the search, since those ARE real turnovers. Rebounds are a special"
        " case: if the action immediately before a shot is the SAME team's own previous shot (a rebound),"
        " that previous shot IS the SCA - labeled simply 'Shot', not the specific outcome type - and the"
        " search stops there. A rebound never gets a second SCA behind it. On the shooting team's own side,"
        " a Tackle or Interception winning the ball still counts as an SCA, but a BallRecovery (picking up a"
        " loose ball rather than winning a contested one) does not - it's skipped over, not counted, so the"
        " search keeps looking further back for the real contributing action. Own goals (a 'Goal' event"
        " carrying WhoScored's 'Own goal' qualifier) are excluded entirely - they carry the same 'isShot'"
        " flag as a real shot in the raw data, but they aren't a shot taken by the team credited with the"
        " goal, so they don't count as a shot and don't get a row on this tab.",
        "",
        "Shot body part (Head / Left Foot / Right Foot / Other) is read directly from WhoScored's own"
        " qualifiers on the shot event.",
        "",
        f"Pitch thirds: Own third = x < {OWN_THIRD_MAX}, Middle third = {OWN_THIRD_MAX} <= x < {MIDDLE_THIRD_MAX},"
        f" Final third = x >= {MIDDLE_THIRD_MAX} (based on the x coordinate only, 0-100 normalized scale,"
        " each team's own attacking direction).",
        "",
        f"Attacking box = inside the opponent's box, assumed {BOX_DEPTH_YD} yards deep and {BOX_WIDTH_YD} yards"
        " wide, centered on the goal.",
        "",
        "Passing tab: per-player totals, one table per team. Passes Completed/Attempted/Forward and Headed"
        " count ALL pass attempts (open play, corners, free kicks, throw-ins, goal kicks, keeper throws)"
        " regardless of outcome, except Passes Completed which requires a successful outcome. Passes Forward"
        " = attempts where the end X coordinate is greater than the start X coordinate (net progress toward"
        " the opponent's goal along that axis), regardless of whether the pass succeeded. Headed = pass"
        " attempts carrying WhoScored's own 'HeadPass' qualifier, regardless of outcome. Crosses Attempted/"
        "Completed are OPEN PLAY only - passes with the 'Cross' qualifier, excluding any also tagged as a"
        " corner or free kick (direct or indirect). Passes into Final 1/3 and Passes into the Box are"
        " completed, open-play passes only (excluding corners, free kicks, throw-ins, goal kicks, and"
        " keeper throws) that start outside that zone and end inside it. Progressive Passes uses the same"
        " rolling-window definition described earlier in this Notes tab. Shot Assists is read directly from WhoScored's"
        " own 'ShotAssist' qualifier on the pass event (the pass immediately preceding a shot attempt),"
        " rather than re-derived from the Shot Creating Actions search. SCA is the combined total of times a"
        " player appears as either the SCA1 or SCA2 contributing action on the Shot Creating Actions tab"
        " (same search/logic as that tab, just totalled per player).",
        "",
        "Totals tab: one row per team, rolling up the same numbers already computed on the other tabs -"
        " Shots is a count of shot events (same rows as the Shot Creating Actions tab, so own goals are"
        " excluded here too - see that tab's note); Touches/thirds/Attacking Box and the carry columns come from the"
        " Touches tab's team summary; Total Passes is the sum of each player's Passes Attempted; Crosses"
        " Attempted/Completed, Passes into Final Third, and Passes into Opponent's Box are summed from the"
        " Passing tab (open-play only, as defined there); Total SCA is the sum of each player's SCA column;"
        " Progressive Passes is the sum of each player's Progressive Passes column from the Passing tab.",
        "",
        "10+ Pass Sequences / Avg Passes per Sequence: possession sequences are built using the same"
        " windowed 'possession chain' algorithm found in the open-source WhoScored report notebooks behind"
        " insight90.streamlit.app - rather than ending a sequence the instant the opponent's team touches"
        " the ball, possession only flips to the other team once there's a sustained run of their events"
        " within a sliding window (a single missed tackle, aerial, or loose touch doesn't end a sequence,"
        " but a real spell of opponent possession does); sequences also always restart at goals and at the"
        " start of a new period. The window size and required run-length were tuned by testing several"
        " combinations against insight90's published numbers for this exact match (25 sequences of 10+"
        " passes at a 9-pass average for Man Utd, 10 sequences at a 6-pass average for Crystal Palace) -"
        " the closest match found (24/13 sequences, 9.69/5.97 average) is close but not an exact"
        " reproduction, since the exact source implementation isn't public. 'Passes' counts ALL pass"
        " attempts in a sequence (not just completed ones), which was also the closest-matching definition."
        " 10+ Pass Sequences counts sequences with at least 10 passes; Avg Passes per Sequence is the mean"
        " across ALL of a team's sequences for the match, not just the long ones.",
        "",
        "Field Tilt: each team's share of the two teams' combined touches in their attacking (final) third,"
        " as a percentage - touches_in_final_third / (both teams' touches_in_final_third). Matches"
        " insight90's published Field Tilt for this match almost exactly.",
        "",
        f"PPDA (Passes Per Defensive Action): the number of passes a team allows its opponent to attempt,"
        f" divided by that team's own defensive actions (Tackle, Interception, Foul, Challenge - any"
        f" outcome; Aerials and Clearances excluded), counting only actions/passes outside each team's own"
        f" deepest {PPDA_CUTOFF}% of the pitch. Lower means more aggressive pressing. Tuned against"
        " insight90's published PPDA for this match; close (6.6 / 15.3 vs a published 6.92 / 15.42) but not"
        " an exact reproduction, since the source implementation isn't public.",
        "",
        "Tackles / Successful Tackles / Interceptions / Blocked Passes / Blocked Shots: exact counts read"
        " directly from WhoScored's own discrete event types - Tackle, Interception, and BlockedPass events"
        " (Blocked Passes is the blocking team's own event, not the passer's), and shots carrying a"
        " 'Blocked' qualifier, credited to the OPPONENT of the team that took the shot (i.e. the side that"
        " did the blocking). These verified as EXACT matches against insight90's published numbers for this"
        " match.",
        "",
        "Defensive Action Height (m): the MEAN x position (converted to metres) of a team's Tackle,"
        " Interception, Clearance, Ball Recovery, and Challenge events, excluding that team's own"
        " goalkeeper - unlike the counts above, this one is tuned/approximate rather than an exact event"
        " count, since the source's precise definition isn't public. Chosen by grid-searching formula"
        " variants (event types, mean vs median, with/without the goalkeeper) against TWO real matches at"
        " once rather than one - the earlier median-based, Aerial-included, GK-included version fit one"
        " match well but fell apart on a second (see diagnose_def_action_height.py for the full search).",
        "",
        "Passes Received / Progressive Passes Received (Touches tab): WhoScored/Opta pass events have no"
        " explicit receiver field, so the receiver is inferred the same way most open-source pass-network"
        " scripts do it - the player of the very next event in the full chronological log, provided that"
        " next event belongs to the SAME team (otherwise no receiver is recorded). Progressive Passes"
        " Received uses the same logic, restricted to passes already flagged progressive.",
        "",
        "Shot Creating Actions tab: one table per team (own goals already excluded - see the Totals"
        " tab note above) rather than one combined table with a Team column.",
        "",
        "Defensive Actions tab: per-player totals, one table per team, for every player who appears"
        " anywhere in the match. Tackles and Interceptions are direct event counts; Passes Blocked counts"
        " 'BlockedPass' events by their own player (the blocker, not the passer). Shots Blocked has no"
        " direct player field in the data - a blocked shot only carries a 'Blocked' qualifier on the"
        " shooter's own event - but WhoScored pairs every blocked shot with a 'Save' event on the very next"
        " row, credited to the defender or keeper who actually stopped it; that player is used as the"
        " blocker. Verified to sum to the same exact team totals as the Totals tab's Blocked Shots column.",
        "",
        "Defensive Action Location tab: the same per-player inclusiveness as the Defensive Actions tab, but"
        " Tackles/Interceptions/Passes Blocked/Ball Recoveries are each split into three columns - Own Third,"
        " Middle Third, Final Third - using the exact same pitch-third rule as the Touches tab (x < "
        f"{OWN_THIRD_MAX} / < {MIDDLE_THIRD_MAX} / else, based on the action's own x coordinate). Ball"
        " Recoveries is a new metric for this tab (WhoScored's 'BallRecovery' event type) not present on the"
        " Defensive Actions tab.",
        "",
        "Possession % (Totals tab): a passes-based possession proxy - each team's total passes attempted"
        " divided by both teams' combined passes attempted. Close to insight90's published possession"
        " (60.8% / 39.2% here vs a published 61% / 39%).",
        "",
        "Corners (Totals tab): count of Pass events carrying the 'CornerTaken' qualifier. Exact match"
        " against insight90's published corner counts for this match.",
        "",
        "Against tab: the mirror image of the Totals tab - same columns, but each row shows the OPPONENT's"
        " numbers rather than the team's own. E.g. the 'vs Man Utd' row's Shots value is Crystal Palace's"
        " total shots, since that's how many shots were taken against Man Utd. Only supports the standard"
        " 2-team match case.",
        "",
        "Passing Pairs tab: every distinct passer -> receiver combination (completed passes only, any pass"
        " type - open play, corners, free kicks, etc.), with a count of how many times that exact pair"
        " happened, one table per team. Receiver uses the same next-event, same-team heuristic as the"
        " Passes Received column on the Touches tab - WhoScored/Opta pass events carry no explicit receiver"
        " field. Sorted by count, descending, within each team.",
        "",
        "Shot Pairs tab: every distinct passer -> shot-taker combination, with a count of how many times"
        " that exact pair happened, one table per team. The passer is whoever played the SCA1 action (the"
        " very first Shot Creating Actions column) for that shot, and only counts when that immediate"
        " preceding action was itself a pass (any type). Shots whose SCA1 was a take-on, a duel, a rebound"
        " off a blocked/saved shot, or a loose ball with no such preceding pass aren't a passer -> shooter"
        " pair and are excluded from this tab. Sorted by count, descending, within each team.",
        "",
        "On-Off tab: per-player totals, one table per team, for every player who appears anywhere in the"
        " match - Shots, Total Touches, Own/Middle/Final Third Touches, Attacking Box Touches, Carries into"
        " Final Third, Carries into Box, Passes into Final 1/3, Passes into the Box, Crosses Attempted, and"
        " Crosses Completed - each counted only during the minutes that player was actually on the pitch,"
        " split into a block of all 'For' columns (that player's own team) followed by a block of all"
        " 'Against' columns (the opponent). Interceptions, Blocked Passes, and Blocked Shots are also"
        " included but 'For' only (no 'Against' pair) - they're already inherently one-sided defensive"
        " actions, so an 'Against' version would be a different stat (how often the OPPONENT blocked/"
        " intercepted this team) rather than a natural counterpart. Interceptions/Blocked Passes are direct"
        " event-type counts; Blocked Shots uses the same definition as the Totals tab's own Blocked Shots"
        " column - the OPPONENT's shot attempts carrying a 'Blocked' qualifier, credited to this team as the"
        " side that did the blocking. A player's on-pitch window runs from kickoff (or their own"
        " SubstitutionOn time, if they came on as a sub) to the final whistle (or their own SubstitutionOff"
        " time, if they were subbed off) - a substituted player isn't credited for the exact minute they"
        " left, since that minute belongs to whoever replaced them. Third/Attacking Box/Carries use the same"
        " definitions as the Touches tab; Passes into Final 1/3 / Passes into the Box / Crosses Attempted /"
        " Crosses Completed use the same OPEN PLAY ONLY definitions as the Passing tab (excludes corners and"
        " free kicks) - all applied to each side's own attacking direction, so e.g. Final Third Against means"
        " the OPPONENT was in THEIR OWN attacking final third (deep in this player's own defensive third),"
        " not the same physical end of the pitch as Final Third For. A red/second-yellow card isn't accounted"
        " for - a sent-off player's window still runs to the final whistle rather than ending at the card.",
    ]
    for i, line in enumerate(notes, start=1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(name='Arial', bold=(i == 1))
    ws.column_dimensions['A'].width = 120

    return wb


# ============================================================
# 7. MAIN
# ============================================================
def sanitize_filename(name):
    return re.sub(r'[^A-Za-z0-9]+', '', name) if name else 'Unknown'


def main():
    if len(sys.argv) < 2:
        print('Usage: python whoscored_report.py "<whoscored match centre url>"')
        sys.exit(1)

    url = sys.argv[1]
    df, match_info = scrape_match(url)
    home_name = match_info.get('home_name')
    away_name = match_info.get('away_name')
    print(f"Scraped {len(df)} events. Teams: {home_name} vs {away_name}")

    print("Computing progressive passes...")
    _, player_totals, team_totals, progressive_received = compute_progressive_passes(df)

    print("Computing passes received...")
    passes_received = compute_passes_received(df)

    print("Computing passing pairs...")
    passing_pairs = compute_passing_pairs(df)

    print("Computing carries...")
    team_carries, player_carries, carries_df = compute_carries(df)

    print("Computing shot-creating actions...")
    sca_out = compute_sca(df)

    print("Computing shot pairs...")
    shot_pairs = compute_shot_pairs(sca_out)

    print("Computing touches...")
    team_summary, player_third = compute_touches(df, team_carries, player_carries,
                                                  passes_received, progressive_received)

    print("Computing passing...")
    passing_out = compute_passing(df, player_totals, sca_out)

    print("Computing possession sequences...")
    chains_df, team_sequences = compute_sequences(df)

    print("Computing field tilt and PPDA...")
    field_tilt = compute_field_tilt(team_summary)
    ppda = compute_ppda(df)

    print("Computing defensive stats...")
    defensive_stats = compute_defensive_stats(df)
    defensive_actions = compute_defensive_actions(df)
    defensive_action_location = compute_defensive_action_location(df)

    print("Computing corners...")
    corners = compute_corners(df)

    print("Computing totals...")
    totals_out = compute_totals(team_summary, team_totals, passing_out, sca_out, chains_df, team_sequences,
                                 field_tilt, ppda, defensive_stats, corners, home_name, away_name)

    print("Computing against totals...")
    against_totals = compute_against_totals(totals_out)

    print("Computing On/Off splits...")
    player_windows = extract_player_windows(df)
    on_off = compute_on_off(df, player_windows, carries_df)

    print("Building workbook...")
    wb = build_workbook(sca_out, team_summary, player_third, passing_out, totals_out, defensive_actions,
                         defensive_action_location, passing_pairs, home_name, away_name, against_totals,
                         shot_pairs, on_off)

    filename = f"{sanitize_filename(home_name)}_vs_{sanitize_filename(away_name)}.xlsx"
    wb.save(filename)
    print(f"Saved: {os.path.abspath(filename)}")


if __name__ == '__main__':
    main()
